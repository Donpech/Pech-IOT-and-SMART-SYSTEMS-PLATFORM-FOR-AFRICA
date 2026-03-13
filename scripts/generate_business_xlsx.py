#!/usr/bin/env python3
"""
PECH Group Holdings Ltd — Business & Operations Documents Excel Generator
Generates branded, fillable Excel spreadsheets for all remaining documents.

Usage: python3 scripts/generate_business_xlsx.py
"""

import os, sys

# Import shared branding helpers from the employment script
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_employment_xlsx import (
    add_branded_header, add_section_header, add_table_headers,
    add_form_field, add_footer, set_column_widths, add_print_setup,
    FILL_LIGHT_ORANGE, FILL_WHITE, FILL_LIGHT_BLUE, FILL_DARK_BLUE,
    FONT_INPUT, FONT_NORMAL, FONT_LABEL, THIN_BORDER, ALIGN_LEFT,
    OUTPUT_DIR, ensure_output_dir,
)
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation


def _tracker_rows(ws, row, max_col, count, validations=None):
    """Add empty tracker rows with alternating colors and optional validations."""
    for i in range(count):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_LEFT
        ws.cell(row=row, column=1).value = i + 1
        ws.row_dimensions[row].height = 22
        if validations:
            for col_idx, dv in validations.items():
                dv.add(ws.cell(row=row, column=col_idx))
        row += 1
    return row


def _make_workbook(title, subtitle, headers, widths, row_count=30, validations_spec=None, max_col=None):
    """Generic tracker workbook builder."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit
    mc = max_col or len(headers)
    set_column_widths(ws, widths)
    row = add_branded_header(ws, title.upper(), subtitle, mc)
    row = add_section_header(ws, row, f"{title.upper()} REGISTER", mc)
    header_row = row
    row = add_table_headers(ws, row, headers)
    data_start_row = row  # First data row (after table headers)

    dvs = {}
    if validations_spec:
        for col_idx, opts in validations_spec.items():
            dv = DataValidation(type="list", formula1=f'"{opts}"')
            ws.add_data_validation(dv)
            dvs[col_idx] = dv

    row = _tracker_rows(ws, row, mc, row_count, dvs)
    add_footer(ws, row, mc)

    # Add auto-filter on the header row for easy sorting/filtering
    from openpyxl.utils import get_column_letter
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(mc)}{data_start_row + row_count - 1}"

    add_print_setup(ws, freeze_row=data_start_row)
    return wb


def _make_form_sheet(wb, sheet_title, doc_title, subtitle, sections, max_col=6):
    """Add a fillable form sheet to a workbook."""
    ws = wb.create_sheet(sheet_title[:31])
    set_column_widths(ws, [4, 30, 32, 4, 30, 32][:max_col])
    row = add_branded_header(ws, doc_title, subtitle, max_col)
    form_start_row = row  # First form row after header

    for section_name, fields in sections:
        row = add_section_header(ws, row, section_name, max_col)
        for label in fields:
            row = add_form_field(ws, row, label, 2, 3, 1, max_col)
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws, freeze_row=form_start_row)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# 1. SALES INVOICE
# ═══════════════════════════════════════════════════════════════════════════════
def create_sales_invoice():
    wb = _make_workbook(
        "Sales Invoice", "Tax Invoice Register — FIRS Compliant",
        ["#", "Invoice No.", "Date", "Client Name", "Client TIN", "Description",
         "Subtotal (NGN)", "VAT 7.5%", "WHT", "Discount", "Total Due",
         "Payment Terms", "Paid", "Status"],
        [5, 16, 14, 22, 16, 24, 16, 14, 14, 14, 16, 14, 14, 14],
        row_count=40,
        validations_spec={13: "Yes,No,Partial", 14: "Draft,Sent,Paid,Overdue,Cancelled"}
    )
    _make_form_sheet(wb, "Invoice Form (Fillable)", "SALES INVOICE (TAX INVOICE)", "FIRS Compliant", [
        ("INVOICE DETAILS", ["Invoice Number:", "Invoice Date:", "Due Date:", "PO Reference:", "Payment Terms:"]),
        ("BUYER / CLIENT", ["Client Name:", "Company:", "Address:", "City/State:", "Phone:", "Email:", "TIN:"]),
        ("ITEMS / SERVICES", [f"Item {i} — Description:" for i in range(1, 11)]),
        ("SUMMARY", ["Subtotal (NGN):", "VAT @ 7.5%:", "WHT (if applicable):", "Discount:", "TOTAL DUE:", "Amount in Words:"]),
        ("BANK DETAILS (NGN)", ["Bank Name:", "Account Name:", "Account Number:", "Sort Code:"]),
        ("AUTHORIZATION", ["Prepared By:", "Signature:", "Date:", "Approved By:", "Signature:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_SALES_INVOICE.xlsx"))
    print("  [OK] Sales Invoice")


# ═══════════════════════════════════════════════════════════════════════════════
# 2. PROFORMA INVOICE
# ═══════════════════════════════════════════════════════════════════════════════
def create_proforma_invoice():
    wb = _make_workbook(
        "Proforma Invoice", "Pre-Sales Quotation Register",
        ["#", "Proforma No.", "Date", "Client Name", "Description",
         "Subtotal (NGN)", "VAT 7.5%", "Total", "Valid Until",
         "Converted to Invoice", "Invoice No.", "Status"],
        [5, 16, 14, 22, 24, 16, 14, 16, 14, 14, 16, 14],
        validations_spec={10: "Yes,No", 12: "Draft,Sent,Accepted,Expired,Rejected"}
    )
    _make_form_sheet(wb, "Proforma Form (Fillable)", "PROFORMA INVOICE", "Pre-Sale Estimate", [
        ("PROFORMA DETAILS", ["Proforma No.:", "Date:", "Valid Until:", "Project Ref.:"]),
        ("CLIENT INFORMATION", ["Client Name:", "Company:", "Address:", "Phone:", "Email:"]),
        ("ITEMS / SERVICES", [f"Item {i} — Description / Qty / Price:" for i in range(1, 9)]),
        ("SUMMARY", ["Subtotal:", "VAT @ 7.5%:", "Discount:", "GRAND TOTAL:", "Amount in Words:"]),
        ("AUTHORIZATION", ["Prepared By:", "Signature:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_PROFORMA_INVOICE.xlsx"))
    print("  [OK] Proforma Invoice")


# ═══════════════════════════════════════════════════════════════════════════════
# 3. QUOTATION
# ═══════════════════════════════════════════════════════════════════════════════
def create_quotation():
    wb = _make_workbook(
        "Quotation", "Client Quotation Register",
        ["#", "Quotation No.", "Date", "Client Name", "Project/Ref",
         "Items Summary", "Subtotal (NGN)", "VAT", "Discount", "Grand Total",
         "Valid Until", "Accepted", "PO Received", "Status"],
        [5, 16, 14, 22, 18, 24, 16, 14, 14, 16, 14, 12, 12, 14],
        row_count=35,
        validations_spec={12: "Yes,No,Pending", 13: "Yes,No,Pending", 14: "Draft,Sent,Accepted,Expired,Rejected,Revised"}
    )
    _make_form_sheet(wb, "Quotation Form (Fillable)", "QUOTATION", "Valid for 30 days", [
        ("QUOTATION DETAILS", ["Quotation Ref. No.:", "Date:", "Valid Until:", "Project/Reference:"]),
        ("CLIENT INFORMATION", ["Company/Client Name:", "Attention:", "Address:", "City/State:", "Phone:", "Email:"]),
        ("QUOTED ITEMS", [f"Item {i} — Desc / Qty / Unit / Price:" for i in range(1, 9)]),
        ("OPTIONAL ADD-ONS", ["Add-on 1:", "Add-on 2:", "Add-on 3:"]),
        ("PRICING SUMMARY", ["Subtotal (Quoted):", "Subtotal (Optional):", "Discount %:", "Net Amount:", "VAT @ 7.5%:", "GRAND TOTAL:", "Amount in Words:"]),
        ("DELIVERY TIMELINE", ["Phase 1:", "Phase 2:", "Phase 3:", "Total Timeline:"]),
        ("PAYMENT TERMS", ["Milestone 1 — %:", "Milestone 2 — %:", "Milestone 3 — %:", "Payment Method:"]),
        ("AUTHORIZATION", ["Prepared By:", "Title:", "Signature:", "Date:", "Client Acceptance Signature:", "Client Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_QUOTATION.xlsx"))
    print("  [OK] Quotation")


# ═══════════════════════════════════════════════════════════════════════════════
# 4. PURCHASE ORDER
# ═══════════════════════════════════════════════════════════════════════════════
def create_purchase_order():
    wb = _make_workbook(
        "Purchase Order", "Procurement Tracking Register",
        ["#", "PO Number", "Date", "Supplier", "Description",
         "Subtotal (NGN)", "VAT", "Shipping", "Grand Total",
         "Delivery Date", "Received", "Invoice Matched", "Paid", "Status"],
        [5, 16, 14, 22, 24, 16, 14, 14, 16, 14, 12, 12, 12, 14],
        row_count=40,
        validations_spec={11: "Yes,No,Partial", 12: "Yes,No,Pending", 13: "Yes,No,Partial", 14: "Draft,Sent,Confirmed,Delivered,Cancelled,Closed"}
    )
    _make_form_sheet(wb, "PO Form (Fillable)", "PURCHASE ORDER", "Procurement", [
        ("PO DETAILS", ["PO Number:", "Date Issued:", "Delivery Date Required:", "Payment Terms:", "Shipping Method:", "Project/Job Ref.:"]),
        ("SUPPLIER INFORMATION", ["Supplier Name:", "Contact Person:", "Address:", "City/State:", "Phone:", "Email:", "TIN:", "CAC Reg. No.:", "Bank Name:", "Account Number:"]),
        ("ORDER ITEMS", [f"Item {i} — Desc / Qty / Unit / Price:" for i in range(1, 11)]),
        ("TOTALS", ["Subtotal (NGN):", "VAT @ 7.5%:", "Shipping/Delivery:", "GRAND TOTAL:"]),
        ("DELIVERY INSTRUCTIONS", ["Delivery Address:", "Contact at Delivery:", "Phone:", "Special Instructions:"]),
        ("APPROVAL", ["Requested By:", "Date:", "Approved By:", "Signature:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_PURCHASE_ORDER.xlsx"))
    print("  [OK] Purchase Order")


# ═══════════════════════════════════════════════════════════════════════════════
# 5. DELIVERY NOTE
# ═══════════════════════════════════════════════════════════════════════════════
def create_delivery_note():
    wb = _make_workbook(
        "Delivery Note", "Outbound Delivery Register",
        ["#", "DN Number", "Date", "Client Name", "Invoice Ref",
         "PO Ref", "Items Summary", "Qty Total", "Packages",
         "Delivered By", "Received By", "Condition", "Status"],
        [5, 16, 14, 22, 16, 16, 24, 12, 12, 18, 18, 14, 14],
        row_count=40,
        validations_spec={12: "Good,Damaged,Partial", 13: "Prepared,In Transit,Delivered,Returned"}
    )
    _make_form_sheet(wb, "DN Form (Fillable)", "DELIVERY NOTE", "Goods Dispatch", [
        ("DELIVERY NOTE DETAILS", ["DN Number:", "Date:", "Invoice Ref:", "PO Ref:", "Quotation Ref:"]),
        ("DELIVERY FROM", ["Company:", "Contact Person:", "Phone:", "Email:"]),
        ("DELIVER TO", ["Client/Company:", "Attention:", "Address:", "City/State:", "Phone:", "Email:"]),
        ("ITEMS DELIVERED", [f"Item {i} — Desc / Qty / Serial No.:" for i in range(1, 9)]),
        ("DISPATCH", ["Total Packages:", "Special Handling:", "Dispatched By:", "Dispatch Date/Time:"]),
        ("RECEIPT CONFIRMATION", ["Received By (Name):", "Signature:", "Date/Time:", "Condition on Receipt:", "Comments:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_DELIVERY_NOTE.xlsx"))
    print("  [OK] Delivery Note")


# ═══════════════════════════════════════════════════════════════════════════════
# 6. GOODS RECEIVED NOTE (GRN)
# ═══════════════════════════════════════════════════════════════════════════════
def create_grn():
    wb = _make_workbook(
        "Goods Received Note", "Inbound Goods Register",
        ["#", "GRN Number", "Date", "Supplier", "PO Ref",
         "DN/Invoice Ref", "Items Summary", "Qty Ordered", "Qty Received",
         "Qty Rejected", "Condition", "Inspected By", "Stored Location", "Status"],
        [5, 16, 14, 22, 16, 16, 22, 12, 12, 12, 14, 18, 16, 14],
        row_count=40,
        validations_spec={11: "Good,Damaged,Mixed", 14: "Received,Inspecting,Accepted,Rejected,Partial"}
    )
    _make_form_sheet(wb, "GRN Form (Fillable)", "GOODS RECEIVED NOTE", "Inbound Inspection", [
        ("GRN DETAILS", ["GRN Number:", "Date Received:", "PO Reference:", "Supplier DN/Invoice:"]),
        ("SUPPLIER INFO", ["Supplier Name:", "Contact:", "Phone:"]),
        ("ITEMS RECEIVED", [f"Item {i} — Desc / Qty Ordered / Qty Received / Condition:" for i in range(1, 9)]),
        ("INSPECTION", ["Inspected By:", "Date:", "Overall Condition:", "Discrepancies:", "Action Required:"]),
        ("STORAGE", ["Stored At:", "Bin/Location:", "Entered in System By:", "Date:"]),
        ("APPROVAL", ["Warehouse Manager:", "Signature:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_GOODS_RECEIVED_NOTE.xlsx"))
    print("  [OK] Goods Received Note")


# ═══════════════════════════════════════════════════════════════════════════════
# 7. CREDIT NOTE
# ═══════════════════════════════════════════════════════════════════════════════
def create_credit_note():
    wb = _make_workbook(
        "Credit Note", "Credit Note Register",
        ["#", "CN Number", "Date", "Client Name", "Original Invoice",
         "Reason", "Credit Amount", "VAT Adj.", "Total Credit",
         "Applied to Invoice", "Refunded", "Status"],
        [5, 16, 14, 22, 16, 20, 16, 14, 16, 16, 12, 14],
        validations_spec={11: "Yes,No,N/A", 12: "Draft,Issued,Applied,Refunded,Cancelled"}
    )
    _make_form_sheet(wb, "CN Form (Fillable)", "CREDIT NOTE", "Adjustment Document", [
        ("CREDIT NOTE DETAILS", ["CN Number:", "Date:", "Original Invoice No.:", "Original Invoice Date:", "Original Amount:", "PO Reference:"]),
        ("CLIENT INFO", ["Client Name:", "Company:", "Address:", "Phone:", "Email:", "TIN:"]),
        ("REASON FOR CREDIT", ["Reason (select):", "Detailed Explanation:"]),
        ("CREDIT ITEMS", [f"Item {i} — Desc / Qty / Credit Amount:" for i in range(1, 6)]),
        ("SUMMARY", ["Subtotal Credit:", "VAT Adjustment (7.5%):", "TOTAL CREDIT:", "Amount in Words:"]),
        ("AUTHORIZATION", ["Prepared By:", "Date:", "Approved By:", "Signature:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_CREDIT_NOTE.xlsx"))
    print("  [OK] Credit Note")


# ═══════════════════════════════════════════════════════════════════════════════
# 8. DEBIT NOTE
# ═══════════════════════════════════════════════════════════════════════════════
def create_debit_note():
    wb = _make_workbook(
        "Debit Note", "Debit Note Register",
        ["#", "DN Number", "Date", "Supplier/Client", "Original Invoice",
         "Reason", "Debit Amount", "VAT Adj.", "Total Debit",
         "Payment Received", "Status"],
        [5, 16, 14, 22, 16, 20, 16, 14, 16, 14, 14],
        validations_spec={10: "Yes,No,Pending", 11: "Draft,Issued,Paid,Outstanding,Cancelled"}
    )
    _make_form_sheet(wb, "Debit Note Form (Fillable)", "DEBIT NOTE", "Charge Adjustment", [
        ("DEBIT NOTE DETAILS", ["Debit Note No.:", "Date:", "Original Invoice:", "Original Date:", "PO Reference:"]),
        ("RECIPIENT INFO", ["Name:", "Company:", "Address:", "Phone:", "Email:", "TIN:"]),
        ("REASON FOR DEBIT", ["Reason:", "Detailed Explanation:"]),
        ("DEBIT ITEMS", [f"Item {i} — Desc / Qty / Debit Amount:" for i in range(1, 6)]),
        ("SUMMARY", ["Subtotal Debit:", "VAT (7.5%):", "TOTAL DEBIT:", "Amount in Words:"]),
        ("AUTHORIZATION", ["Prepared By:", "Date:", "Approved By:", "Signature:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_DEBIT_NOTE.xlsx"))
    print("  [OK] Debit Note")


# ═══════════════════════════════════════════════════════════════════════════════
# 9. PAYMENT VOUCHER
# ═══════════════════════════════════════════════════════════════════════════════
def create_payment_voucher():
    wb = _make_workbook(
        "Payment Voucher", "Payment Voucher Register",
        ["#", "Voucher No.", "Date", "Payee Name", "Department",
         "Description", "Gross Amt (NGN)", "WHT", "VAT", "Deductions",
         "Net Amount", "Payment Method", "Cheque/Ref No.", "Status"],
        [5, 16, 14, 22, 16, 22, 16, 14, 14, 14, 16, 14, 16, 14],
        row_count=40,
        validations_spec={12: "Bank Transfer,Cheque,Cash,Mobile Money", 14: "Draft,Pending Approval,Approved,Paid,Rejected"}
    )
    _make_form_sheet(wb, "PV Form (Fillable)", "PAYMENT VOUCHER", "Finance Department", [
        ("VOUCHER DETAILS", ["Voucher Number:", "Voucher Date:", "Payment Period:", "Department:"]),
        ("PAYEE DETAILS", ["Payee Name:", "Company:", "Address:", "Phone:", "Email:", "TIN:"]),
        ("PAYEE BANK", ["Bank Name:", "Account Name:", "Account Number:", "Sort Code:"]),
        ("PAYMENT ITEMS", [f"Item {i} — Description / GL Code / Amount:" for i in range(1, 7)]),
        ("SUMMARY", ["Gross Amount:", "Less: WHT:", "Less: VAT:", "Less: Other Deductions:", "NET PAYABLE:", "Amount in Words:"]),
        ("APPROVAL", ["Prepared By:", "Date:", "Checked By:", "Date:", "Approved By:", "Signature:", "Date:", "Received By:", "Signature:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_PAYMENT_VOUCHER.xlsx"))
    print("  [OK] Payment Voucher")


# ═══════════════════════════════════════════════════════════════════════════════
# 10. PETTY CASH VOUCHER
# ═══════════════════════════════════════════════════════════════════════════════
def create_petty_cash():
    wb = _make_workbook(
        "Petty Cash Voucher", "Petty Cash Disbursement Register",
        ["#", "Voucher No.", "Date", "Requested By", "Department",
         "Purpose", "Amount (NGN)", "Receipt Attached", "Approved By",
         "Disbursed By", "Status"],
        [5, 16, 14, 20, 16, 24, 16, 14, 18, 18, 14],
        row_count=50,
        validations_spec={8: "Yes,No", 11: "Pending,Approved,Disbursed,Rejected,Retired"}
    )
    _make_form_sheet(wb, "PCV Form (Fillable)", "PETTY CASH VOUCHER", "Max NGN 50,000", [
        ("VOUCHER DETAILS", ["Voucher Number:", "Date:", "Department:"]),
        ("REQUESTER", ["Requested By:", "Position:", "Phone:"]),
        ("EXPENSE DETAILS", ["Purpose:", f"Item 1:", "Item 2:", "Item 3:", "Item 4:", "Item 5:"]),
        ("AMOUNTS", ["Total Amount Requested:", "Amount in Words:", "Receipt(s) Attached:"]),
        ("APPROVAL", ["Approved By:", "Signature:", "Date:", "Disbursed By:", "Signature:", "Date:", "Received By:", "Signature:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_PETTY_CASH_VOUCHER.xlsx"))
    print("  [OK] Petty Cash Voucher")


# ═══════════════════════════════════════════════════════════════════════════════
# 11. PAYMENT RECEIPT
# ═══════════════════════════════════════════════════════════════════════════════
def create_payment_receipt():
    wb = _make_workbook(
        "Payment Receipt", "Payment Receipt Register",
        ["#", "Receipt No.", "Date", "Received From", "Company",
         "Description", "Amount (NGN)", "Payment Method", "Cheque/Ref",
         "Invoice Ref", "Issued By", "Status"],
        [5, 16, 14, 22, 20, 24, 16, 16, 16, 16, 18, 14],
        row_count=40,
        validations_spec={8: "Bank Transfer,Cash,Cheque,POS,Mobile Money", 12: "Issued,Voided,Duplicate"}
    )
    _make_form_sheet(wb, "Receipt Form (Fillable)", "PAYMENT RECEIPT", "Proof of Payment", [
        ("RECEIPT DETAILS", ["Receipt Number:", "Date:", "Invoice Reference:"]),
        ("RECEIVED FROM", ["Name:", "Company:", "Address:", "Phone:", "Email:"]),
        ("PAYMENT INFO", ["Amount (NGN):", "Amount in Words:", "Payment Method:", "Cheque/Transfer Ref:", "Bank:", "Description:"]),
        ("AUTHORIZATION", ["Received By:", "Signature:", "Date:", "Cashier/Finance Officer:", "Stamp:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_PAYMENT_RECEIPT.xlsx"))
    print("  [OK] Payment Receipt")


# ═══════════════════════════════════════════════════════════════════════════════
# 12. EXPENSE REPORT
# ═══════════════════════════════════════════════════════════════════════════════
def create_expense_report():
    wb = _make_workbook(
        "Expense Report", "Employee Expense Claims Register",
        ["#", "Report No.", "Date", "Employee Name", "Department",
         "Purpose/Trip", "Period", "Total Claimed", "Advance Received",
         "Net Due/Refund", "Receipts", "Approved By", "Paid", "Status"],
        [5, 16, 14, 20, 16, 22, 14, 16, 16, 16, 12, 18, 12, 14],
        row_count=40,
        validations_spec={11: "All,Partial,None", 13: "Yes,No,Pending", 14: "Draft,Submitted,Approved,Paid,Rejected"}
    )
    _make_form_sheet(wb, "Expense Form (Fillable)", "EXPENSE REPORT", "Attach all receipts", [
        ("EMPLOYEE DETAILS", ["Employee Name:", "Employee ID:", "Department:", "Position:", "Phone:", "Line Manager:"]),
        ("EXPENSE PERIOD", ["Start Date:", "End Date:", "Days:", "Purpose:", "Destination:", "Project/Client Ref:"]),
        ("PRE-APPROVAL", ["Pre-Approved:", "Approval Ref:", "Budget (NGN):", "Advance Received:", "Advance Receipt Ref:"]),
        ("ITEMIZED EXPENSES", [f"Item {i} — Date / Description / Category / Amount / Receipt (Y/N):" for i in range(1, 16)]),
        ("CATEGORY SUMMARY", ["Transport:", "Accommodation:", "Meals:", "Communication:", "Entertainment:", "Office Supplies:", "Registration:", "Miscellaneous:", "Other:"]),
        ("FINANCIAL SUMMARY", ["Total Claimed:", "Less: Advance:", "NET DUE TO EMPLOYEE:", "Or REFUND DUE TO COMPANY:"]),
        ("APPROVAL", ["Employee Signature:", "Date:", "Line Manager:", "Date:", "Finance:", "Date:", "MD/CEO (if >NGN 100K):", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_EXPENSE_REPORT.xlsx"))
    print("  [OK] Expense Report")


# ═══════════════════════════════════════════════════════════════════════════════
# 13. ASSET REGISTER
# ═══════════════════════════════════════════════════════════════════════════════
def create_asset_register():
    wb = _make_workbook(
        "Asset Register", "Company Asset Tracking",
        ["#", "Asset Tag", "Date Acquired", "Category", "Description",
         "Make/Brand", "Model", "Serial No.", "Purchase Price (NGN)",
         "Current Value", "Location", "Assigned To", "Condition",
         "Warranty Expiry", "Status"],
        [5, 18, 14, 16, 22, 16, 16, 18, 16, 16, 16, 18, 14, 14, 14],
        row_count=50,
        validations_spec={4: "IT Equipment,Furniture,Vehicle,Tool/Machinery,Electronics,Solar,IoT Device,Other",
                          13: "New,Good,Fair,Poor,Decommissioned",
                          15: "In Use,In Storage,Under Repair,Disposed,Lost/Stolen"}
    )
    _make_form_sheet(wb, "Asset Form (Fillable)", "ASSET REGISTER / TRACKING FORM", "Fixed Asset Management", [
        ("ASSET IDENTIFICATION", ["Asset Tag Number:", "Date Acquired:", "Asset Category:"]),
        ("ASSET DESCRIPTION", ["Description:", "Make/Brand:", "Model Number:", "Serial Number:", "Color:", "Specifications:"]),
        ("ACQUISITION DETAILS", ["Supplier Name:", "Supplier Contact:", "PO Reference:", "Invoice Reference:", "Purchase Price (NGN):", "Current Value (NGN):", "Warranty Period:", "Warranty Expiry:", "Insurance Policy:"]),
        ("LOCATION & ASSIGNMENT", ["Current Location:", "Department:", "Assigned To:", "Employee ID:", "Date Assigned:", "Condition:"]),
        ("TRANSFER HISTORY", ["Transfer 1 — From / To / Date / Reason:", "Transfer 2 — From / To / Date / Reason:", "Transfer 3 — From / To / Date / Reason:"]),
        ("MAINTENANCE HISTORY", ["Service 1 — Date / Description / Cost:", "Service 2 — Date / Description / Cost:", "Service 3 — Date / Description / Cost:", "Total Maintenance Cost:"]),
        ("DISPOSAL", ["Disposal Date:", "Disposal Method:", "Disposal Value:", "Authorized By:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_ASSET_REGISTER_FORM.xlsx"))
    print("  [OK] Asset Register")


# ═══════════════════════════════════════════════════════════════════════════════
# 14. STOCK REQUISITION
# ═══════════════════════════════════════════════════════════════════════════════
def create_stock_requisition():
    wb = _make_workbook(
        "Stock Requisition", "Internal Stock Request Register",
        ["#", "Requisition No.", "Date", "Requested By", "Department",
         "Items Description", "Qty Requested", "Qty Issued", "Purpose",
         "Approved By", "Issued By", "Date Issued", "Status"],
        [5, 16, 14, 20, 16, 24, 14, 14, 20, 18, 18, 14, 14],
        row_count=40,
        validations_spec={13: "Pending,Approved,Partially Issued,Issued,Rejected,Cancelled"}
    )
    _make_form_sheet(wb, "Requisition Form (Fillable)", "STOCK REQUISITION FORM", "Internal Materials Request", [
        ("REQUISITION DETAILS", ["Requisition Number:", "Date:", "Required By Date:", "Priority:"]),
        ("REQUESTER", ["Requested By:", "Department:", "Position:", "Phone:"]),
        ("ITEMS REQUESTED", [f"Item {i} — Description / Qty / Unit / Purpose:" for i in range(1, 11)]),
        ("APPROVAL", ["Line Manager:", "Signature:", "Date:", "Store Manager:", "Signature:", "Date:"]),
        ("ISSUE DETAILS", ["Issued By:", "Date Issued:", "Received By:", "Signature:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_STOCK_REQUISITION_FORM.xlsx"))
    print("  [OK] Stock Requisition")


# ═══════════════════════════════════════════════════════════════════════════════
# 15. WORK ORDER
# ═══════════════════════════════════════════════════════════════════════════════
def create_work_order():
    wb = _make_workbook(
        "Work Order", "Work Order Tracking Register",
        ["#", "WO Number", "Date", "Priority", "Work Type",
         "Client/Site", "Description", "Assigned Team", "Start Date",
         "Due Date", "Completed Date", "Cost (NGN)", "Quality Score", "Status"],
        [5, 16, 14, 12, 16, 22, 24, 18, 14, 14, 14, 16, 12, 14],
        row_count=40,
        validations_spec={4: "Urgent,High,Medium,Low",
                          5: "Installation,Maintenance,Repair,Inspection,Other",
                          14: "Open,Assigned,In Progress,Completed,On Hold,Cancelled"}
    )
    _make_form_sheet(wb, "WO Form (Fillable)", "WORK ORDER", "Field Operations", [
        ("WORK ORDER DETAILS", ["WO Number:", "Date Issued:", "Priority:", "Work Type:", "Project/Job Ref:"]),
        ("REQUESTER", ["Requested By:", "Department:", "Phone:", "Email:"]),
        ("CLIENT / SITE", ["Client Name:", "Site Address:", "City/State:", "Site Contact:", "Site Phone:"]),
        ("SCOPE OF WORK", ["Description:", "Task 1:", "Task 2:", "Task 3:", "Task 4:", "Task 5:"]),
        ("SAFETY REQUIREMENTS", ["Work Permit Required:", "PPE Required:", "Site Restrictions:"]),
        ("MATERIALS NEEDED", [f"Material {i} — Description / Qty:" for i in range(1, 6)]),
        ("TEAM ASSIGNMENT", ["Team Lead:", "Team Members:", "Start Date:", "Due Date:", "Estimated Hours:"]),
        ("COMPLETION", ["Completed Date:", "Actual Hours:", "Total Cost:", "Quality Score:", "Client Sign-Off:", "Remarks:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_WORK_ORDER.xlsx"))
    print("  [OK] Work Order")


# ═══════════════════════════════════════════════════════════════════════════════
# 16. WAYBILL
# ═══════════════════════════════════════════════════════════════════════════════
def create_waybill():
    wb = _make_workbook(
        "Waybill", "Goods Transportation Register",
        ["#", "Waybill No.", "Date", "Origin", "Destination",
         "Consignor", "Consignee", "Driver Name", "Vehicle Reg.",
         "Items Summary", "Total Weight", "Declared Value", "Delivered", "Status"],
        [5, 16, 14, 18, 18, 20, 20, 18, 16, 22, 14, 16, 12, 14],
        row_count=40,
        validations_spec={13: "Yes,No,Partial", 14: "Prepared,In Transit,Delivered,Returned,Lost"}
    )
    _make_form_sheet(wb, "Waybill Form (Fillable)", "WAYBILL", "Transportation Document", [
        ("WAYBILL DETAILS", ["Waybill Number:", "Date:", "Origin:", "Destination:"]),
        ("CONSIGNOR (SENDER)", ["Company/Name:", "Address:", "City/State:", "Contact:", "Phone:", "Email:"]),
        ("CONSIGNEE (RECEIVER)", ["Company/Name:", "Address:", "City/State:", "Contact:", "Phone:", "Email:"]),
        ("VEHICLE / TRANSPORT", ["Vehicle Type:", "Registration No.:", "Driver Name:", "Driver Phone:", "Driver Licence:", "Transport Company:", "Estimated Transit:"]),
        ("GOODS DESCRIPTION", [f"Item {i} — Desc / Qty / Weight / Dimensions / Value:" for i in range(1, 7)]),
        ("TOTALS", ["Total Packages:", "Total Weight:", "Total Declared Value:", "Freight Charge:", "Insurance:"]),
        ("SIGNATURES", ["Consignor Signature:", "Date:", "Driver Signature:", "Date:", "Consignee Signature:", "Date:", "Condition on Receipt:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_WAYBILL.xlsx"))
    print("  [OK] Waybill")


# ═══════════════════════════════════════════════════════════════════════════════
# 17. VISITOR LOG
# ═══════════════════════════════════════════════════════════════════════════════
def create_visitor_log():
    wb = _make_workbook(
        "Visitor Log", "Daily Visitor Registration",
        ["#", "Date", "Visitor Name", "Company", "Phone",
         "ID Type", "ID Number", "Purpose", "Person to See",
         "Department", "Time In", "Time Out", "Badge No.",
         "Vehicle Reg.", "Items Declared", "Status"],
        [5, 14, 20, 20, 16, 14, 16, 18, 18, 16, 12, 12, 12, 14, 16, 14],
        row_count=60,
        validations_spec={6: "National ID,Driver Licence,Passport,Voter Card",
                          8: "Meeting,Delivery,Maintenance,Interview,Tour,Vendor,Other",
                          16: "Checked In,On Premises,Checked Out"}
    )
    wb.save(os.path.join(OUTPUT_DIR, "PECH_VISITOR_LOG.xlsx"))
    print("  [OK] Visitor Log")


# ═══════════════════════════════════════════════════════════════════════════════
# 18. INCIDENT REPORT
# ═══════════════════════════════════════════════════════════════════════════════
def create_incident_report():
    wb = _make_workbook(
        "Incident Report", "Safety & Incident Tracking",
        ["#", "Incident Ref.", "Date", "Time", "Location",
         "Type", "Severity", "Persons Involved", "Description Summary",
         "Immediate Action", "Root Cause", "Reported By", "Investigated By", "Status"],
        [5, 16, 14, 12, 18, 16, 14, 20, 24, 22, 22, 18, 18, 14],
        row_count=30,
        validations_spec={6: "Workplace Injury,Property Damage,Security Breach,Near Miss,Fire/Electrical,Theft/Loss,Vehicle Accident,Environmental,Other",
                          7: "Critical,Major,Minor,Negligible",
                          14: "Reported,Under Investigation,Resolved,Closed,Escalated"}
    )
    _make_form_sheet(wb, "Incident Form (Fillable)", "INCIDENT REPORT FORM", "H&S Compliance", [
        ("INCIDENT DETAILS", ["Incident Ref. No.:", "Date of Incident:", "Time:", "Date Reported:", "Location:", "Specific Area/Room:"]),
        ("CLASSIFICATION", ["Incident Type:", "Severity Level:"]),
        ("PERSONS INVOLVED", ["Person 1 — Name / Role / Dept / Injury:", "Person 2 — Name / Role / Dept / Injury:", "Person 3 — Name / Role / Dept / Injury:"]),
        ("WITNESSES", ["Witness 1 — Name / Phone / Email:", "Witness 2 — Name / Phone / Email:"]),
        ("DESCRIPTION", ["Detailed Account:", "Events Leading Up:", "Contributing Factors:"]),
        ("IMMEDIATE ACTIONS", ["First Aid Given:", "Emergency Services Called:", "Area Secured:", "Immediate Actions Taken:"]),
        ("INVESTIGATION", ["Root Cause:", "Corrective Actions:", "Preventive Measures:", "Responsible Person:", "Completion Deadline:"]),
        ("SIGN-OFF", ["Reported By:", "Date:", "Investigated By:", "Date:", "HSE Officer:", "Date:", "Management:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_INCIDENT_REPORT_FORM.xlsx"))
    print("  [OK] Incident Report")


# ═══════════════════════════════════════════════════════════════════════════════
# 19. INSTALLATION COMPLETION CERTIFICATE
# ═══════════════════════════════════════════════════════════════════════════════
def create_installation_cert():
    wb = _make_workbook(
        "Installation Completion", "Installation Certificate Register",
        ["#", "Certificate No.", "Date", "Client Name", "Site Address",
         "Work Order Ref.", "Installer Name", "Devices Installed",
         "Tests Passed", "Client Signed", "Photos Attached", "Warranty Start", "Status"],
        [5, 16, 14, 22, 22, 16, 20, 18, 14, 14, 14, 14, 14],
        row_count=30,
        validations_spec={9: "All Pass,Partial,Fail", 10: "Yes,No,Pending", 11: "Yes,No",
                          13: "Draft,Completed,Submitted,Accepted,Disputed"}
    )
    _make_form_sheet(wb, "Certificate Form (Fillable)", "INSTALLATION COMPLETION CERTIFICATE", "IoT/Solar Installation", [
        ("CERTIFICATE DETAILS", ["Certificate No.:", "Date:", "Work Order Ref.:", "Project Ref.:"]),
        ("CLIENT INFO", ["Client Name:", "Company:", "Site Address:", "City/State:", "Phone:", "Email:"]),
        ("INSTALLATION DETAILS", ["Installer Name:", "Installer ID:", "Installation Date:", "Completion Date:"]),
        ("DEVICES INSTALLED", [f"Device {i} — Type / Model / Serial / Location:" for i in range(1, 7)]),
        ("TESTING & COMMISSIONING", ["Power Test:", "Connectivity Test:", "Data Transmission:", "Sensor Calibration:", "Platform Registration:", "All Tests Pass:"]),
        ("SIGN-OFF", ["Installer Signature:", "Date:", "Client Signature:", "Client Name:", "Date:", "PECH Supervisor:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_INSTALLATION_COMPLETION_CERTIFICATE.xlsx"))
    print("  [OK] Installation Completion Certificate")


# ═══════════════════════════════════════════════════════════════════════════════
# 20. JOB COMPLETION REPORT
# ═══════════════════════════════════════════════════════════════════════════════
def create_job_completion():
    wb = _make_workbook(
        "Job Completion Report", "Project/Job Completion Tracking",
        ["#", "Job Ref.", "Date Completed", "Client Name", "Description",
         "Assigned To", "Start Date", "End Date", "Budget (NGN)",
         "Actual Cost", "Quality Score", "Client Feedback", "Invoice Raised", "Status"],
        [5, 16, 14, 22, 24, 18, 14, 14, 16, 16, 12, 16, 14, 14],
        row_count=30,
        validations_spec={13: "Yes,No,Pending", 14: "Completed,Accepted,Disputed,Rework,Closed"}
    )
    _make_form_sheet(wb, "JCR Form (Fillable)", "JOB COMPLETION REPORT", "Post-Job Documentation", [
        ("JOB DETAILS", ["Job Reference:", "Work Order No.:", "Project Name:", "Client Name:", "Site Address:"]),
        ("TIMELINE", ["Start Date:", "Planned End Date:", "Actual End Date:", "Days Variance:"]),
        ("WORK SUMMARY", ["Scope of Work:", "Tasks Completed:", "Deviations:", "Additional Work:"]),
        ("MATERIALS USED", [f"Material {i} — Description / Qty / Cost:" for i in range(1, 6)]),
        ("FINANCIALS", ["Budget:", "Actual Cost:", "Variance:", "Invoice Amount:", "Invoice No.:"]),
        ("QUALITY", ["Quality Score (1-10):", "Client Satisfaction:", "Issues/Defects:", "Warranty Terms:"]),
        ("SIGN-OFF", ["Completed By:", "Date:", "Verified By:", "Date:", "Client Acceptance:", "Date:"]),
    ])
    wb.save(os.path.join(OUTPUT_DIR, "PECH_JOB_COMPLETION_REPORT.xlsx"))
    print("  [OK] Job Completion Report")


# ═══════════════════════════════════════════════════════════════════════════════
# 21. HACKATHON RULES & GUIDELINES
# ═══════════════════════════════════════════════════════════════════════════════
def create_hackathon():
    wb = _make_workbook(
        "Hackathon Participants", "PECH Hackathon Event Tracking",
        ["#", "Participant/Team", "Team Size", "Track",
         "Contact Email", "Phone", "Registration Date",
         "Project Name", "Demo Link", "R1 Score", "R2 Score",
         "Final Score", "Prize", "Status"],
        [5, 22, 10, 20, 22, 16, 14, 22, 22, 12, 12, 12, 16, 14],
        row_count=50,
        validations_spec={4: "IoT & Smart Devices,Marketplace & Commerce,Logistics & Delivery,Creator & Content,Open Innovation",
                          14: "Registered,Submitted,Judging,Winner,Runner-Up,Completed,Disqualified"}
    )
    wb.save(os.path.join(OUTPUT_DIR, "PECH_HACKATHON_RULES_AND_GUIDELINES.xlsx"))
    print("  [OK] Hackathon Rules & Guidelines")


# ═══════════════════════════════════════════════════════════════════════════════
# 22. MEDIA & PUBLIC COMMUNICATIONS POLICY
# ═══════════════════════════════════════════════════════════════════════════════
def create_media_policy():
    wb = _make_workbook(
        "Media Comms Register", "Public Communications Approval Tracking",
        ["#", "Request Date", "Employee Name", "Department",
         "Communication Type", "Platform/Outlet", "Topic/Subject",
         "Pre-Approved Content", "CEO Approval", "Date Approved",
         "Published Date", "Link/Reference", "Status"],
        [5, 14, 20, 16, 18, 18, 24, 14, 14, 14, 14, 22, 14],
        row_count=30,
        validations_spec={5: "Interview,Press Statement,Social Media,Blog Post,Conference,Article,Podcast,Other",
                          8: "Yes,No,Pending", 9: "Yes,No,Pending",
                          13: "Requested,Approved,Published,Denied,Retracted"}
    )
    wb.save(os.path.join(OUTPUT_DIR, "PECH_MEDIA_AND_PUBLIC_COMMUNICATIONS_POLICY.xlsx"))
    print("  [OK] Media & Public Communications Policy")


# ═══════════════════════════════════════════════════════════════════════════════
# 23. SCHEDULE B TEMPLATES (Role KPIs)
# ═══════════════════════════════════════════════════════════════════════════════
def create_schedule_b():
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule B — All Roles"
    max_col = 10
    set_column_widths(ws, [5, 14, 28, 18, 24, 24, 24, 24, 24, 14])

    row = add_branded_header(ws, "SCHEDULE B: JOB DESCRIPTION & KPIs", "Attached to Employment Contracts", max_col)
    row = add_section_header(ws, row, "ALL ROLE KPI TEMPLATES (30 ROLES)", max_col)

    headers = ["#", "Role ID", "Job Title", "Department",
               "KPI 1", "KPI 2", "KPI 3", "KPI 4", "KPI 5", "Review"]
    row = add_table_headers(ws, row, headers)

    roles_kpis = [
        ("PECH-ENG-CTO-001", "CTO", "Engineering", "System uptime >99.5%", "Sprint velocity +10%/quarter", "Hire pipeline filled", "Security: 0 critical vulns", "Tech debt <15%"),
        ("PECH-ENG-FS-001", "Full-Stack Engineer", "Engineering", "Feature delivery on sprint", "Code review participation", "Bug fix turnaround <48h", "Unit test coverage >80%", "Documentation updated"),
        ("PECH-ENG-FE-001", "Frontend Engineer", "Engineering", "Page load <2s", "Lighthouse score >90", "Cross-browser compat.", "Accessibility WCAG AA", "Component library growth"),
        ("PECH-ENG-BE-001", "Backend Engineer", "Engineering", "API response <200ms p95", "Zero downtime deploys", "Error rate <0.1%", "Database query optimized", "API documentation 100%"),
        ("PECH-ENG-IOT-001", "IoT Engineer", "Engineering", "Device connectivity >98%", "Firmware OTA success >99%", "Sensor accuracy ±2%", "Battery life targets met", "Certification compliance"),
        ("PECH-ENG-MOB-001", "Mobile Engineer", "Engineering", "App store rating >4.5", "Crash-free rate >99.5%", "App load <3s", "Feature parity web/mobile", "User retention +5%/month"),
        ("PECH-ENG-DEV-001", "DevOps Engineer", "Engineering", "CI/CD pipeline <10min", "Infrastructure cost -10%", "Incident response <15min", "Backup/restore tested monthly", "Security patches <24h"),
        ("PECH-ENG-DATA-001", "Sr Data Engineer", "Engineering", "Data pipeline uptime >99%", "Query performance SLA met", "ML model accuracy targets", "Data quality score >95%", "Cost per inference -15%"),
        ("PECH-ENG-DATA-002", "Data Analyst", "Engineering", "Weekly reports on time", "Dashboard uptime >99%", "Insight accuracy verified", "Stakeholder satisfaction >4/5", "Self-service adoption +20%"),
        ("PECH-ENG-QA-001", "QA Engineer", "Engineering", "Test coverage >85%", "Bug escape rate <5%", "Regression suite pass rate", "Test automation +20%/qtr", "Release go/no-go accuracy"),
        ("PECH-DES-UI-001", "UI/UX Designer", "Design", "Design system components", "User task success >90%", "Prototype turnaround <3d", "Usability test score >80", "Brand consistency 100%"),
        ("PECH-DES-GRA-001", "Graphic Designer", "Design", "Asset turnaround <2d", "Brand guideline adherence", "Social media engagement", "Print quality 100%", "Template library growth"),
        ("PECH-BIZ-PM-001", "Payments PM", "Product & Business", "PSSP licence milestones", "Transaction success >99.5%", "Feature adoption rate", "Stakeholder NPS >8", "Regulatory compliance 100%"),
        ("PECH-BIZ-BD-001", "BD Lead", "Product & Business", "Pipeline value target", "Conversion rate >15%", "Partnership agreements", "Revenue growth targets", "Client retention >90%"),
        ("PECH-BIZ-MKT-001", "Market Ops Manager", "Product & Business", "Vendor onboarding target", "GMV growth monthly", "Market coverage expansion", "Vendor satisfaction >4/5", "Operational cost ratio"),
        ("PECH-BIZ-COMP-001", "Compliance Officer", "Product & Business", "Regulatory filings on time", "Audit findings resolved", "Policy update cadence", "Training completion 100%", "Zero compliance breaches"),
        ("PECH-OPS-LOG-001", "Logistics Coordinator", "Operations", "Delivery SLA >95%", "Cost per delivery -10%", "Route optimization savings", "Damage rate <1%", "Driver utilization >85%"),
        ("PECH-OPS-CS-001", "CS Lead", "Operations", "First response <1h", "Resolution time <24h", "CSAT score >4.5/5", "Ticket backlog <20", "Team training hours/month"),
        ("PECH-OPS-CS-002", "CS Rep", "Operations", "Tickets resolved/day >15", "CSAT per agent >4/5", "First contact resolution >70%", "Knowledge base updates", "Escalation rate <10%"),
        ("PECH-OPS-FIN-001", "Finance Officer", "Operations", "Books closed by 5th", "Reconciliation accuracy", "Expense report turnaround", "Audit-ready at all times", "Cash flow forecast accuracy"),
    ]

    dv_review = DataValidation(type="list", formula1='"Quarterly,Monthly,Bi-Annual,Annual"')
    ws.add_data_validation(dv_review)

    for idx, (role_id, title, dept, *kpis) in enumerate(roles_kpis):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_NORMAL
            cell.fill = FILL_LIGHT_ORANGE if idx % 2 else FILL_WHITE
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_LEFT
        ws.cell(row=row, column=1).value = idx + 1
        ws.cell(row=row, column=2).value = role_id
        ws.cell(row=row, column=3).value = title
        ws.cell(row=row, column=4).value = dept
        for k, kpi in enumerate(kpis):
            ws.cell(row=row, column=5 + k).value = kpi
        ws.cell(row=row, column=10).value = "Quarterly"
        dv_review.add(ws.cell(row=row, column=10))
        ws.row_dimensions[row].height = 28
        row += 1

    # Empty rows for remaining roles
    for i in range(10):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = len(roles_kpis) + i + 1
        dv_review.add(ws.cell(row=row, column=10))
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_SCHEDULE_B_TEMPLATES.xlsx"))
    print("  [OK] Schedule B Templates")


# ═══════════════════════════════════════════════════════════════════════════════
# UPDATED MASTER DASHBOARD (replaces the old one)
# ═══════════════════════════════════════════════════════════════════════════════
def update_master_dashboard():
    """Create comprehensive master dashboard covering ALL documents."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Dashboard"
    max_col = 8
    set_column_widths(ws, [5, 36, 20, 14, 14, 14, 14, 26])

    row = add_branded_header(ws, "PECH MASTER DOCUMENT DASHBOARD", "All Employment + Business Documents", max_col)

    row = add_section_header(ws, row, "COMPLETE DOCUMENT INDEX (38 Excel Files)", max_col)
    headers = ["#", "Document Name", "Category", "Total Records", "Active", "Pending", "Completed", "Google Sheets Link"]
    row = add_table_headers(ws, row, headers)

    all_docs = [
        # Employment / HR (15)
        ("Full-Time Employment Contract", "Employment"),
        ("Contract Worker Agreement", "Employment"),
        ("Internship Agreement", "Employment"),
        ("Office Marketer Agreement", "Employment"),
        ("Commission Marketer Agreement", "Employment"),
        ("Installer Agreement", "Employment"),
        ("API Developer Agreement", "Employment"),
        ("Guarantor / Surety Form", "Employment"),
        ("Non-Disclosure Agreement", "Employment"),
        ("Candidate Application Form", "HR Form"),
        ("Interview Process & Checklist", "HR Form"),
        ("Leave Request Form", "HR Form"),
        ("Staff ID Card Request", "HR Form"),
        ("Job Requirements Handbook", "HR Reference"),
        ("Schedule B: KPI Templates", "HR Reference"),
        # Finance (8)
        ("Sales Invoice (Tax Invoice)", "Finance"),
        ("Proforma Invoice", "Finance"),
        ("Quotation", "Finance"),
        ("Purchase Order", "Finance"),
        ("Credit Note", "Finance"),
        ("Debit Note", "Finance"),
        ("Payment Voucher", "Finance"),
        ("Payment Receipt", "Finance"),
        # Operations (8)
        ("Expense Report", "Operations"),
        ("Petty Cash Voucher", "Operations"),
        ("Delivery Note", "Operations"),
        ("Goods Received Note", "Operations"),
        ("Stock Requisition Form", "Operations"),
        ("Waybill", "Operations"),
        ("Work Order", "Operations"),
        ("Job Completion Report", "Operations"),
        # Facilities & Safety (3)
        ("Asset Register / Tracking Form", "Assets"),
        ("Visitor Log", "Facilities"),
        ("Incident Report Form", "Safety"),
        # Policy & Events (3)
        ("Media & Public Communications", "Policy"),
        ("Hackathon Rules & Guidelines", "Events"),
        ("Installation Completion Certificate", "Operations"),
    ]

    for idx, (name, cat) in enumerate(all_docs):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_NORMAL
            cell.fill = FILL_LIGHT_ORANGE if idx % 2 else FILL_WHITE
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_LEFT
        ws.cell(row=row, column=1).value = idx + 1
        ws.cell(row=row, column=2).value = name
        ws.cell(row=row, column=3).value = cat
        ws.cell(row=row, column=8).value = "(paste Google Sheets URL)"
        ws.cell(row=row, column=8).font = FONT_INPUT
        row += 1

    row += 2
    row = add_section_header(ws, row, "DOCUMENT CATEGORIES SUMMARY", max_col)
    cat_headers = ["#", "Category", "Document Count", "Description", "", "", "", ""]
    row = add_table_headers(ws, row, cat_headers)

    cats = [
        ("Employment Contracts", "9", "All employment and contractor agreements"),
        ("HR Forms", "6", "Application, interview, leave, ID, handbook, KPIs"),
        ("Finance", "8", "Invoices, POs, quotations, vouchers, receipts, notes"),
        ("Operations", "8", "Delivery, GRN, stock, waybill, work orders, jobs"),
        ("Assets / Facilities / Safety", "3", "Asset register, visitor log, incident reports"),
        ("Policy & Events", "2", "Media policy, hackathon tracker"),
        ("TOTAL", "36", "Complete organizational document suite"),
    ]
    for idx, (cat, count, desc) in enumerate(cats):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=10, bold=(cat == "TOTAL"), color="333333")
            cell.fill = FILL_LIGHT_BLUE if cat == "TOTAL" else (FILL_LIGHT_ORANGE if idx % 2 else FILL_WHITE)
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = idx + 1
        ws.cell(row=row, column=2).value = cat
        ws.cell(row=row, column=3).value = count
        ws.cell(row=row, column=4).value = desc
        row += 1

    row += 2
    row = add_section_header(ws, row, "GOOGLE SHEETS SETUP INSTRUCTIONS", max_col)
    instructions = [
        "1. Open Google Drive (drive.google.com)",
        "2. Click '+ New' > 'File upload' > Select one or more .xlsx files",
        "3. Double-click the uploaded file to open in Google Sheets",
        "4. Google Sheets will automatically convert the file",
        "5. Click 'File' > 'Save as Google Sheets' for native version",
        "6. Share via 'Share' button with team members",
        "7. Set permissions: 'Editor' for relevant team, 'Viewer' for others",
        "8. Copy the URL and paste in the 'Google Sheets Link' column above",
        "",
        "TIP: All dropdown validations, colors, and formatting are preserved!",
        "TIP: Use 'File > Make a copy' to create new templates from filled ones.",
        "TIP: Upload all files to a shared 'PECH Documents' folder in Google Drive.",
    ]
    for instr in instructions:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        cell = ws.cell(row=row, column=1)
        cell.value = f"  {instr}"
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = 20
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_MASTER_DOCUMENT_DASHBOARD.xlsx"))
    print("  [OK] Master Document Dashboard (updated)")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    ensure_output_dir()

    print("\n" + "=" * 70)
    print("  PECH Group Holdings Ltd — Business Documents Excel Generator")
    print("  Generating branded .xlsx for all remaining documents...")
    print("=" * 70 + "\n")

    # Finance Documents
    create_sales_invoice()
    create_proforma_invoice()
    create_quotation()
    create_purchase_order()
    create_credit_note()
    create_debit_note()
    create_payment_voucher()
    create_petty_cash()
    create_payment_receipt()

    # Operations Documents
    create_expense_report()
    create_delivery_note()
    create_grn()
    create_stock_requisition()
    create_work_order()
    create_waybill()
    create_job_completion()
    create_installation_cert()

    # Assets / Facilities / Safety
    create_asset_register()
    create_visitor_log()
    create_incident_report()

    # Contracts / Policy
    create_hackathon()
    create_media_policy()
    create_schedule_b()

    # Updated Master Dashboard
    update_master_dashboard()

    print("\n" + "=" * 70)
    print(f"  SUCCESS! 24 additional Excel files generated in: {OUTPUT_DIR}/")
    print("=" * 70)

    total = len([f for f in os.listdir(OUTPUT_DIR) if f.endswith(".xlsx")])
    print(f"\n  TOTAL files in directory: {total} Excel workbooks")
    print("\n  All files:")
    for f in sorted(os.listdir(OUTPUT_DIR)):
        if f.endswith(".xlsx"):
            size = os.path.getsize(os.path.join(OUTPUT_DIR, f))
            print(f"    - {f} ({size / 1024:.1f} KB)")
    print()


if __name__ == "__main__":
    main()
