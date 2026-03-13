#!/usr/bin/env python3
"""
PECH Group Holdings Ltd — Employment Documents Excel (.xlsx) Generator
Generates branded, fillable Excel spreadsheets for all employment documents.
Compatible with Google Sheets import for cloud-based sharing and tracking.

Usage: python3 scripts/generate_employment_xlsx.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── PECH Brand Colors ────────────────────────────────────────────────────────
PECH_SKY_BLUE = "00BFFF"
PECH_DARK_BLUE = "0099CC"
PECH_ORANGE = "F5A623"
PECH_DARK_ORANGE = "E08A00"
PECH_DARK_BG = "1B2838"
PECH_WHITE = "FFFFFF"
PECH_LIGHT_BLUE = "E6F7FF"
PECH_LIGHT_ORANGE = "FFF5E6"
PECH_LIGHT_GRAY = "F5F5F5"

# ─── Reusable Styles ─────────────────────────────────────────────────────────
FILL_DARK_BG = PatternFill(start_color=PECH_DARK_BG, end_color=PECH_DARK_BG, fill_type="solid")
FILL_BLUE = PatternFill(start_color=PECH_SKY_BLUE, end_color=PECH_SKY_BLUE, fill_type="solid")
FILL_DARK_BLUE = PatternFill(start_color=PECH_DARK_BLUE, end_color=PECH_DARK_BLUE, fill_type="solid")
FILL_ORANGE = PatternFill(start_color=PECH_ORANGE, end_color=PECH_ORANGE, fill_type="solid")
FILL_DARK_ORANGE = PatternFill(start_color=PECH_DARK_ORANGE, end_color=PECH_DARK_ORANGE, fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(start_color=PECH_LIGHT_BLUE, end_color=PECH_LIGHT_BLUE, fill_type="solid")
FILL_LIGHT_ORANGE = PatternFill(start_color=PECH_LIGHT_ORANGE, end_color=PECH_LIGHT_ORANGE, fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color=PECH_LIGHT_GRAY, end_color=PECH_LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=PECH_WHITE, end_color=PECH_WHITE, fill_type="solid")

FONT_TITLE = Font(name="Calibri", size=18, bold=True, color=PECH_WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, bold=True, color=PECH_WHITE)
FONT_COMPANY = Font(name="Calibri", size=10, italic=True, color=PECH_ORANGE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=PECH_WHITE)
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=PECH_DARK_BLUE)
FONT_LABEL = Font(name="Calibri", size=10, bold=True, color=PECH_DARK_BG)
FONT_NORMAL = Font(name="Calibri", size=10, color="333333")
FONT_INPUT = Font(name="Calibri", size=10, color="0066AA")
FONT_FOOTER = Font(name="Calibri", size=8, italic=True, color="999999")
FONT_SMALL_WHITE = Font(name="Calibri", size=9, color=PECH_WHITE)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

BLUE_BORDER = Border(
    left=Side(style="thin", color=PECH_SKY_BLUE),
    right=Side(style="thin", color=PECH_SKY_BLUE),
    top=Side(style="thin", color=PECH_SKY_BLUE),
    bottom=Side(style="thin", color=PECH_SKY_BLUE),
)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "employment_xlsx")


def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def add_branded_header(ws, title, subtitle, max_col=8):
    """Add a colorful branded header block to a worksheet."""
    # Row 1: Top gradient bar (blue)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = FILL_BLUE
        cell.value = ""
    ws.row_dimensions[1].height = 6

    # Row 2: Dark background with company name
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    cell = ws.cell(row=2, column=1)
    cell.value = "PECH GROUP HOLDINGS LTD"
    cell.font = FONT_TITLE
    cell.fill = FILL_DARK_BG
    cell.alignment = ALIGN_CENTER
    for col in range(2, max_col + 1):
        ws.cell(row=2, column=col).fill = FILL_DARK_BG
    ws.row_dimensions[2].height = 36

    # Row 3: Dark background with subtitle
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    cell = ws.cell(row=3, column=1)
    cell.value = title
    cell.font = FONT_SUBTITLE
    cell.fill = FILL_DARK_BG
    cell.alignment = ALIGN_CENTER
    for col in range(2, max_col + 1):
        ws.cell(row=3, column=col).fill = FILL_DARK_BG
    ws.row_dimensions[3].height = 24

    # Row 4: Orange accent bar
    for col in range(1, max_col + 1):
        cell = ws.cell(row=4, column=col)
        cell.fill = FILL_ORANGE
    ws.row_dimensions[4].height = 4

    # Row 5: Company info line
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=max_col)
    cell = ws.cell(row=5, column=1)
    cell.value = f"Lagos, Nigeria  |  pechgroupholdings.tech  |  {subtitle}  |  CONFIDENTIAL"
    cell.font = FONT_COMPANY
    cell.fill = FILL_DARK_BG
    cell.alignment = ALIGN_CENTER
    for col in range(2, max_col + 1):
        ws.cell(row=5, column=col).fill = FILL_DARK_BG
    ws.row_dimensions[5].height = 20

    # Row 6: Bottom gradient bar (orange to blue)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=6, column=col)
        if col <= max_col // 2:
            cell.fill = FILL_ORANGE
        else:
            cell.fill = FILL_BLUE
    ws.row_dimensions[6].height = 4

    # Row 7: Spacer
    ws.row_dimensions[7].height = 8

    return 8  # Next available row


def add_section_header(ws, row, title, max_col=8):
    """Add a colored section header row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    cell.value = f"  {title}"
    cell.font = Font(name="Calibri", size=11, bold=True, color=PECH_WHITE)
    cell.fill = FILL_DARK_BLUE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, max_col + 1):
        ws.cell(row=row, column=col).fill = FILL_DARK_BLUE
    ws.row_dimensions[row].height = 26
    return row + 1


def add_table_headers(ws, row, headers, fills=None):
    """Add styled column headers."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = fills[col_idx - 1] if fills else FILL_DARK_BLUE
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28
    return row + 1


def add_form_field(ws, row, label, col_label=1, col_value=2, value_span=1, max_col=8):
    """Add a label-value form field pair."""
    cell_label = ws.cell(row=row, column=col_label)
    cell_label.value = label
    cell_label.font = FONT_LABEL
    cell_label.fill = FILL_LIGHT_BLUE
    cell_label.alignment = ALIGN_LEFT
    cell_label.border = THIN_BORDER

    if value_span > 1:
        ws.merge_cells(start_row=row, start_column=col_value,
                       end_row=row, end_column=col_value + value_span - 1)
    cell_value = ws.cell(row=row, column=col_value)
    cell_value.font = FONT_INPUT
    cell_value.fill = FILL_WHITE
    cell_value.alignment = ALIGN_LEFT
    cell_value.border = BLUE_BORDER
    # Fill remaining merged cells with border
    for c in range(col_value + 1, col_value + value_span):
        ws.cell(row=row, column=c).border = BLUE_BORDER

    ws.row_dimensions[row].height = 24
    return row + 1


def add_data_row(ws, row, num_cols, alt=False):
    """Add an empty data entry row with alternating colors."""
    fill = FILL_LIGHT_ORANGE if alt else FILL_WHITE
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_INPUT
        cell.fill = fill
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22
    return row + 1


def add_footer(ws, row, max_col=8):
    """Add a branded footer."""
    row += 1
    # Orange bar
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = FILL_ORANGE
    ws.row_dimensions[row].height = 3
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    cell.value = "PECH Group Holdings Ltd  |  Confidential  |  pechgroupholdings.tech  |  Technology & Infrastructure Enablers for People"
    cell.font = FONT_FOOTER
    cell.alignment = ALIGN_CENTER
    for col in range(2, max_col + 1):
        ws.cell(row=row, column=col).fill = FILL_WHITE

    row += 1
    # Blue bar
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = FILL_BLUE
    ws.row_dimensions[row].height = 3
    return row + 1


def set_column_widths(ws, widths):
    """Set column widths from a dict or list."""
    if isinstance(widths, dict):
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width
    elif isinstance(widths, list):
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width


def add_print_setup(ws, freeze_row=None):
    """Configure print settings and freeze panes for better usability."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Freeze panes below header area so data scrolls under the branding
    if freeze_row:
        ws.freeze_panes = ws.cell(row=freeze_row, column=1)
    else:
        # Default: freeze below row 10 (after branded header + table headers)
        ws.freeze_panes = "A10"


# ═══════════════════════════════════════════════════════════════════════════════
# 1. FULL-TIME EMPLOYMENT CONTRACT TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def create_fulltime_contract():
    wb = Workbook()

    # ── Sheet 1: Employee Contract Tracker ──
    ws = wb.active
    ws.title = "FT Contract Tracker"
    max_col = 16
    set_column_widths(ws, [5, 20, 20, 22, 18, 15, 16, 15, 18, 14, 14, 14, 14, 15, 18, 18])

    row = add_branded_header(ws, "FULL-TIME EMPLOYMENT CONTRACT TRACKER", "Track all FT employee contracts", max_col)
    row = add_section_header(ws, row, "EMPLOYEE CONTRACTS REGISTER", max_col)

    headers = ["#", "Employee Name", "NIN", "Job Title", "Department",
               "Start Date", "Salary (NGN/mo)", "Probation End",
               "Reporting To", "Work Location", "Bank Name",
               "Account No.", "Contract Signed", "Guarantors (2)",
               "NDA Signed", "Status"]
    row = add_table_headers(ws, row, headers)

    # Status validation
    dv_status = DataValidation(type="list", formula1='"Active,Probation,Confirmed,Terminated,Resigned,Suspended"')
    dv_status.prompt = "Select status"
    ws.add_data_validation(dv_status)

    dv_yn = DataValidation(type="list", formula1='"Yes,No,Pending"')
    ws.add_data_validation(dv_yn)

    for i in range(30):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.alignment = ALIGN_LEFT
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv_status.add(ws.cell(row=row, column=max_col))
        dv_yn.add(ws.cell(row=row, column=13))
        dv_yn.add(ws.cell(row=row, column=14))
        dv_yn.add(ws.cell(row=row, column=15))
        row += 1

    add_footer(ws, row, max_col)

    # ── Sheet 2: Individual Contract Form ──
    ws2 = wb.create_sheet("Contract Form (Fillable)")
    max_col2 = 6
    set_column_widths(ws2, [4, 28, 30, 4, 28, 30])

    row = add_branded_header(ws2, "FULL-TIME EMPLOYMENT CONTRACT", "Fillable Contract Form", max_col2)
    row = add_section_header(ws2, row, "PARTIES", max_col2)
    row = add_form_field(ws2, row, "Contract Date:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Employee Full Name:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Employee Address:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Employee NIN:", 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "APPOINTMENT & POSITION", max_col2)
    row = add_form_field(ws2, row, "Job Title:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Department:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Reporting To:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Commencement Date:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Place of Work:", 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "REMUNERATION & BENEFITS", max_col2)
    row = add_form_field(ws2, row, "Monthly Salary (NGN):", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Bank Name:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Account Number:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Account Name:", 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "EQUITY (IF APPLICABLE)", max_col2)
    row = add_form_field(ws2, row, "Shares/Options:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "% of Company:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Vesting Period:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Cliff Period:", 2, 3, 1, max_col2)
    row = add_form_field(ws2, row, "Exercise Price:", 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "SIGNATURES", max_col2)
    for label in ["Employer Name:", "Employer Title:", "Employer Signature:", "Employer Date:",
                   "Employee Name:", "Employee Signature:", "Employee Date:",
                   "Witness Name:", "Witness Address:", "Witness Signature:", "Witness Date:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)

    add_footer(ws2, row, max_col2)
    add_print_setup(ws)
    add_print_setup(ws2)

    wb.save(os.path.join(OUTPUT_DIR, "PECH_FULL_TIME_EMPLOYMENT_CONTRACT.xlsx"))
    print("  [OK] Full-Time Employment Contract")


# ═══════════════════════════════════════════════════════════════════════════════
# 2. CONTRACT WORKER AGREEMENT TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def create_contract_worker():
    wb = Workbook()
    ws = wb.active
    ws.title = "Contractor Tracker"
    max_col = 14
    set_column_widths(ws, [5, 20, 20, 22, 15, 15, 16, 14, 14, 15, 14, 14, 18, 14])

    row = add_branded_header(ws, "INDEPENDENT CONTRACTOR AGREEMENT TRACKER", "Track all contractor engagements", max_col)
    row = add_section_header(ws, row, "CONTRACTOR REGISTER", max_col)

    headers = ["#", "Contractor Name", "Company/Business", "Scope of Work",
               "Start Date", "End Date", "Fee (NGN/mo)", "Payment Terms",
               "Contract Ref.", "NDA Signed", "IP Assigned", "Insurance",
               "Status", "Notes"]
    row = add_table_headers(ws, row, headers)

    dv = DataValidation(type="list", formula1='"Active,Completed,Terminated,Pending,Renewed"')
    ws.add_data_validation(dv)
    dv_yn = DataValidation(type="list", formula1='"Yes,No,Pending"')
    ws.add_data_validation(dv_yn)

    for i in range(25):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv.add(ws.cell(row=row, column=13))
        dv_yn.add(ws.cell(row=row, column=10))
        dv_yn.add(ws.cell(row=row, column=11))
        dv_yn.add(ws.cell(row=row, column=12))
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_CONTRACT_WORKER_AGREEMENT.xlsx"))
    print("  [OK] Contract Worker Agreement")


# ═══════════════════════════════════════════════════════════════════════════════
# 3. INTERNSHIP AGREEMENT TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def create_internship():
    wb = Workbook()
    ws = wb.active
    ws.title = "Intern Tracker"
    max_col = 16
    set_column_widths(ws, [5, 20, 18, 18, 20, 15, 15, 14, 16, 18, 18, 14, 14, 14, 14, 16])

    row = add_branded_header(ws, "INTERNSHIP AGREEMENT TRACKER", "Track all intern engagements", max_col)
    row = add_section_header(ws, row, "INTERN REGISTER", max_col)

    headers = ["#", "Intern Name", "DOB", "NIN", "Department/Team",
               "Start Date", "End Date", "Stipend (NGN)", "Supervisor",
               "Institution", "Programme", "Agreement Ref.", "NDA Signed",
               "Mid-Term Eval", "Final Eval", "Status"]
    row = add_table_headers(ws, row, headers)

    dv = DataValidation(type="list", formula1='"Active,Completed,Terminated,Probation,Certificate Issued"')
    ws.add_data_validation(dv)
    dv_eval = DataValidation(type="list", formula1='"Excellent,Good,Satisfactory,Needs Improvement,Not Yet Done"')
    ws.add_data_validation(dv_eval)

    for i in range(20):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv.add(ws.cell(row=row, column=16))
        dv_eval.add(ws.cell(row=row, column=14))
        dv_eval.add(ws.cell(row=row, column=15))
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_INTERNSHIP_AGREEMENT.xlsx"))
    print("  [OK] Internship Agreement")


# ═══════════════════════════════════════════════════════════════════════════════
# 4. OFFICE MARKETER AGREEMENT TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def create_office_marketer():
    wb = Workbook()
    ws = wb.active
    ws.title = "Office Marketer Tracker"
    max_col = 14
    set_column_widths(ws, [5, 20, 18, 15, 15, 16, 16, 14, 14, 16, 14, 14, 14, 16])

    row = add_branded_header(ws, "OFFICE MARKETER EMPLOYMENT TRACKER", "Track office marketer contracts", max_col)
    row = add_section_header(ws, row, "OFFICE MARKETER REGISTER", max_col)

    headers = ["#", "Marketer Name", "NIN", "Start Date", "Probation End",
               "Base Salary (NGN)", "Commission Rate", "Monthly Target",
               "Sales Q1", "Sales Q2", "Sales Q3", "Sales Q4",
               "Contract Ref.", "Status"]
    row = add_table_headers(ws, row, headers)

    dv = DataValidation(type="list", formula1='"Active,Probation,Confirmed,Terminated,Resigned"')
    ws.add_data_validation(dv)

    for i in range(20):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv.add(ws.cell(row=row, column=14))
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_OFFICE_MARKETER_AGREEMENT.xlsx"))
    print("  [OK] Office Marketer Agreement")


# ═══════════════════════════════════════════════════════════════════════════════
# 5. COMMISSION MARKETER AGREEMENT TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def create_commission_marketer():
    wb = Workbook()
    ws = wb.active
    ws.title = "Commission Marketer Tracker"
    max_col = 14
    set_column_widths(ws, [5, 20, 18, 15, 15, 14, 14, 18, 14, 14, 14, 16, 18, 14])

    row = add_branded_header(ws, "COMMISSION MARKETER AGREEMENT TRACKER", "Track commission-based marketers", max_col)
    row = add_section_header(ws, row, "COMMISSION MARKETER REGISTER", max_col)

    headers = ["#", "Marketer Name", "BVN/TIN", "Start Date", "End Date",
               "Base Commission %", "Bonus Tier %", "Territory",
               "Total Sales (NGN)", "Commission Earned", "Paid to Date",
               "Outstanding", "Contract Ref.", "Status"]
    row = add_table_headers(ws, row, headers)

    dv = DataValidation(type="list", formula1='"Active,Suspended,Terminated,Completed"')
    ws.add_data_validation(dv)

    for i in range(25):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv.add(ws.cell(row=row, column=14))
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_COMMISSION_MARKETER_AGREEMENT.xlsx"))
    print("  [OK] Commission Marketer Agreement")


# ═══════════════════════════════════════════════════════════════════════════════
# 6. INSTALLER AGREEMENT TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def create_installer():
    wb = Workbook()
    ws = wb.active
    ws.title = "Installer Tracker"
    max_col = 14
    set_column_widths(ws, [5, 22, 18, 18, 15, 15, 18, 14, 14, 14, 14, 14, 16, 14])

    row = add_branded_header(ws, "CERTIFIED INSTALLER AGREEMENT TRACKER", "Track installer contractors", max_col)
    row = add_section_header(ws, row, "INSTALLER REGISTER", max_col)

    headers = ["#", "Installer Name/Business", "RC/BN Number", "Territory",
               "Start Date", "End Date", "Certification Status",
               "Jobs Completed", "Quality Score", "Insurance Valid",
               "Equipment Issued", "Fee Rate (NGN)", "Contract Ref.", "Status"]
    row = add_table_headers(ws, row, headers)

    dv = DataValidation(type="list", formula1='"Active,Suspended,Terminated,Expired,Renewed"')
    ws.add_data_validation(dv)
    dv_cert = DataValidation(type="list", formula1='"Certified,In Training,Expired,Revoked"')
    ws.add_data_validation(dv_cert)

    for i in range(20):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv.add(ws.cell(row=row, column=14))
        dv_cert.add(ws.cell(row=row, column=7))
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_INSTALLER_AGREEMENT.xlsx"))
    print("  [OK] Installer Agreement")


# ═══════════════════════════════════════════════════════════════════════════════
# 7. GUARANTOR FORM
# ═══════════════════════════════════════════════════════════════════════════════
def create_guarantor_form():
    wb = Workbook()

    # Sheet 1: Guarantor Tracker
    ws = wb.active
    ws.title = "Guarantor Tracker"
    max_col = 14
    set_column_widths(ws, [5, 20, 20, 20, 18, 18, 14, 18, 14, 14, 14, 14, 14, 16])

    row = add_branded_header(ws, "GUARANTOR / SURETY FORM TRACKER", "Track all employee guarantors", max_col)
    row = add_section_header(ws, row, "GUARANTOR REGISTER (2 per employee required)", max_col)

    headers = ["#", "Employee Name", "Guarantor Name", "Guarantor Occupation",
               "Guarantor Phone", "Guarantor Email", "NIN", "Employer",
               "Yrs Known", "Address Verified", "Employment Verified",
               "ID Verified", "Approved", "HR Officer"]
    row = add_table_headers(ws, row, headers)

    dv = DataValidation(type="list", formula1='"Yes,No,Pending"')
    ws.add_data_validation(dv)
    dv_app = DataValidation(type="list", formula1='"Approved,Rejected,Pending"')
    ws.add_data_validation(dv_app)

    for i in range(40):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv.add(ws.cell(row=row, column=10))
        dv.add(ws.cell(row=row, column=11))
        dv.add(ws.cell(row=row, column=12))
        dv_app.add(ws.cell(row=row, column=13))
        row += 1

    add_footer(ws, row, max_col)

    # Sheet 2: Fillable Form
    ws2 = wb.create_sheet("Guarantor Form (Fillable)")
    max_col2 = 6
    set_column_widths(ws2, [4, 30, 32, 4, 30, 32])

    row = add_branded_header(ws2, "GUARANTOR / SURETY FORM", "2 required per employee", max_col2)

    row = add_section_header(ws2, row, "SECTION A: EMPLOYEE INFORMATION", max_col2)
    for label in ["Employee Full Name:", "Position / Job Title:", "Department:",
                   "Employment Start Date:", "Employee ID:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "SECTION B: GUARANTOR PERSONAL INFORMATION", max_col2)
    for label in ["Full Legal Name:", "Date of Birth:", "Residential Address:",
                   "City / State:", "Primary Phone:", "Secondary Phone:",
                   "Email Address:", "Occupation:", "Employer Name:", "Employer Address:",
                   "NIN:", "BVN:", "Gov. ID Type:", "Gov. ID Number:", "ID Expiry Date:",
                   "Relationship to Employee:", "Years Known:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "SECTION I: ATTESTATION", max_col2)
    for label in ["Guarantor Signature:", "Guarantor Name (Print):", "Date:",
                   "Witness Signature:", "Witness Name:", "Witness Address:",
                   "Witness Occupation:", "Witness Date:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "FOR OFFICIAL USE ONLY", max_col2)
    for label in ["Verified By (HR):", "Date Verified:", "Verification Method:",
                   "Address Verified:", "Employment Verified:", "ID Verified:",
                   "Decision:", "Reason (if rejected):", "HR Signature:", "Date:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)

    add_footer(ws2, row, max_col2)
    add_print_setup(ws)
    add_print_setup(ws2)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_GUARANTOR_FORM.xlsx"))
    print("  [OK] Guarantor Form")


# ═══════════════════════════════════════════════════════════════════════════════
# 8. NDA TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def create_nda():
    wb = Workbook()
    ws = wb.active
    ws.title = "NDA Tracker"
    max_col = 12
    set_column_widths(ws, [5, 22, 22, 18, 15, 15, 18, 18, 14, 14, 18, 14])

    row = add_branded_header(ws, "NON-DISCLOSURE AGREEMENT TRACKER", "Track all NDAs", max_col)
    row = add_section_header(ws, row, "NDA REGISTER", max_col)

    headers = ["#", "Receiving Party Name", "Organisation", "NDA Ref. No.",
               "Effective Date", "Expiry Date", "Purpose", "Scope",
               "Signed by PECH", "Signed by Party", "Witness", "Status"]
    row = add_table_headers(ws, row, headers)

    dv = DataValidation(type="list", formula1='"Active,Expired,Terminated,Breached"')
    ws.add_data_validation(dv)

    for i in range(30):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv.add(ws.cell(row=row, column=12))
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_NON_DISCLOSURE_AGREEMENT.xlsx"))
    print("  [OK] Non-Disclosure Agreement")


# ═══════════════════════════════════════════════════════════════════════════════
# 9. API DEVELOPER AGREEMENT TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def create_api_developer():
    wb = Workbook()
    ws = wb.active
    ws.title = "API Developer Tracker"
    max_col = 14
    set_column_widths(ws, [5, 22, 18, 18, 15, 15, 18, 14, 14, 14, 14, 14, 18, 14])

    row = add_branded_header(ws, "API DEVELOPER AGREEMENT TRACKER", "Track external API developers", max_col)
    row = add_section_header(ws, row, "API DEVELOPER REGISTER", max_col)

    headers = ["#", "Developer Name/Company", "Contact Email", "API Key Issued",
               "Start Date", "End Date", "API Access Tier", "Rate Limit",
               "Sandbox Access", "Production Access", "NDA Signed", "Data Compliant",
               "Agreement Ref.", "Status"]
    row = add_table_headers(ws, row, headers)

    dv = DataValidation(type="list", formula1='"Active,Suspended,Revoked,Expired"')
    ws.add_data_validation(dv)
    dv_tier = DataValidation(type="list", formula1='"Free,Basic,Standard,Premium,Enterprise"')
    ws.add_data_validation(dv_tier)

    for i in range(20):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv.add(ws.cell(row=row, column=14))
        dv_tier.add(ws.cell(row=row, column=7))
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_API_DEVELOPER_AGREEMENT.xlsx"))
    print("  [OK] API Developer Agreement")


# ═══════════════════════════════════════════════════════════════════════════════
# 10. CANDIDATE APPLICATION FORM
# ═══════════════════════════════════════════════════════════════════════════════
def create_candidate_application():
    wb = Workbook()

    # Sheet 1: Application Tracker
    ws = wb.active
    ws.title = "Application Tracker"
    max_col = 16
    set_column_widths(ws, [5, 18, 18, 22, 18, 14, 14, 15, 14, 15, 14, 14, 14, 14, 16, 16])

    row = add_branded_header(ws, "CANDIDATE APPLICATION TRACKER", "Track all job applications", max_col)
    row = add_section_header(ws, row, "APPLICATIONS REGISTER", max_col)

    headers = ["#", "First Name", "Surname", "Position Applied", "Department",
               "Emp. Type", "Date Applied", "Expected Salary", "Source",
               "CV Received", "Certificates", "NIN Verified", "Shortlisted",
               "Interview Stage", "Offer Made", "Final Status"]
    row = add_table_headers(ws, row, headers)

    dv_status = DataValidation(type="list", formula1='"Received,Screening,Round 1,Round 2,Round 3,Offered,Hired,Rejected,Waitlisted,Withdrawn"')
    ws.add_data_validation(dv_status)
    dv_dept = DataValidation(type="list", formula1='"Engineering,Design,Product & Business,Operations,Internship"')
    ws.add_data_validation(dv_dept)
    dv_type = DataValidation(type="list", formula1='"Full-Time,Contract,Internship"')
    ws.add_data_validation(dv_type)
    dv_yn = DataValidation(type="list", formula1='"Yes,No,Pending"')
    ws.add_data_validation(dv_yn)

    for i in range(50):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv_dept.add(ws.cell(row=row, column=5))
        dv_type.add(ws.cell(row=row, column=6))
        dv_yn.add(ws.cell(row=row, column=10))
        dv_yn.add(ws.cell(row=row, column=11))
        dv_yn.add(ws.cell(row=row, column=12))
        dv_yn.add(ws.cell(row=row, column=13))
        dv_yn.add(ws.cell(row=row, column=15))
        dv_status.add(ws.cell(row=row, column=16))
        row += 1

    add_footer(ws, row, max_col)

    # Sheet 2: Fillable Application Form
    ws2 = wb.create_sheet("Application Form (Fillable)")
    max_col2 = 6
    set_column_widths(ws2, [4, 30, 30, 4, 30, 30])

    row = add_branded_header(ws2, "CANDIDATE APPLICATION FORM", "NDPA 2023 Compliant", max_col2)

    row = add_section_header(ws2, row, "SECTION 1: PERSONAL INFORMATION", max_col2)
    for label in ["Surname / Last Name:", "First Name:", "Middle Name:",
                   "Date of Birth (DD/MM/YYYY):", "Gender:", "Nationality:",
                   "State of Origin:", "LGA:", "Phone (Primary):", "Phone (Secondary):",
                   "WhatsApp:", "Email:", "Residential Address:", "City:", "State:",
                   "NIN:", "Other ID Type:", "ID Number:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "SECTION 2: POSITION APPLIED FOR", max_col2)
    for label in ["Position / Job Title:", "Department:", "Employment Type:",
                   "Expected Monthly Salary:", "Available Start Date:", "Notice Period:",
                   "How Did You Hear About Us?"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "SECTION 3: EDUCATION", max_col2)
    for label in ["Tertiary Institution 1:", "Qualification:", "Course:", "Year:",
                   "Tertiary Institution 2:", "Qualification:", "Course:", "Year:",
                   "Secondary School:", "Certificate:", "Year:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "SECTION 4: WORK EXPERIENCE (Most Recent)", max_col2)
    for label in ["Company Name:", "Industry:", "Job Title:", "Employment Type:",
                   "Start Date:", "End Date:", "Salary:", "Supervisor:", "Reason for Leaving:",
                   "Key Responsibilities:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "SECTION 6: REFERENCES", max_col2)
    for ref_num in [1, 2]:
        for label in [f"Reference {ref_num} Name:", f"Ref {ref_num} Organisation:",
                      f"Ref {ref_num} Phone:", f"Ref {ref_num} Email:"]:
            row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "SECTION 10: DECLARATION", max_col2)
    for label in ["Applicant Full Name:", "Signature:", "Date:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)

    add_footer(ws2, row, max_col2)
    add_print_setup(ws)
    add_print_setup(ws2)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_CANDIDATE_APPLICATION_FORM.xlsx"))
    print("  [OK] Candidate Application Form")


# ═══════════════════════════════════════════════════════════════════════════════
# 11. INTERVIEW PROCESS & SCORING
# ═══════════════════════════════════════════════════════════════════════════════
def create_interview_checklist():
    wb = Workbook()

    # Sheet 1: Interview Pipeline Tracker
    ws = wb.active
    ws.title = "Interview Pipeline"
    max_col = 16
    set_column_widths(ws, [5, 18, 18, 22, 18, 14, 14, 14, 14, 12, 12, 12, 14, 14, 16, 16])

    row = add_branded_header(ws, "INTERVIEW PROCESS & SCORING TRACKER", "Internal HR Use Only", max_col)
    row = add_section_header(ws, row, "INTERVIEW PIPELINE", max_col)

    headers = ["#", "First Name", "Surname", "Position", "Level",
               "R1 Date", "R1 Score", "R1 Pass", "R2 Date", "R2 Score",
               "R2 Pass", "R3 Date", "R3 Score", "Overall Score",
               "Decision", "Notes"]
    row = add_table_headers(ws, row, headers)

    dv_level = DataValidation(type="list", formula1='"C-Level,Senior,Mid-Level,Junior,Intern,Field Agent"')
    ws.add_data_validation(dv_level)
    dv_pass = DataValidation(type="list", formula1='"Pass,Fail,Pending,N/A"')
    ws.add_data_validation(dv_pass)
    dv_decision = DataValidation(type="list", formula1='"Hire,Reject,Waitlist,Pending"')
    ws.add_data_validation(dv_decision)

    for i in range(40):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv_level.add(ws.cell(row=row, column=5))
        dv_pass.add(ws.cell(row=row, column=8))
        dv_pass.add(ws.cell(row=row, column=11))
        dv_decision.add(ws.cell(row=row, column=15))
        row += 1

    add_footer(ws, row, max_col)

    # Sheet 2: Scoring Rubric
    ws2 = wb.create_sheet("Scoring Rubric")
    max_col2 = 8
    set_column_widths(ws2, [5, 28, 12, 12, 12, 12, 12, 22])

    row = add_branded_header(ws2, "CANDIDATE SCORING RUBRIC", "PECH Interview Scoring System", max_col2)

    row = add_section_header(ws2, row, "SCORING CRITERIA (1-5 Scale)", max_col2)
    headers2 = ["#", "Criteria", "1 - Poor", "2 - Below Avg", "3 - Average", "4 - Good", "5 - Excellent", "Weight"]
    row = add_table_headers(ws2, row, headers2)

    criteria = [
        ("Technical Knowledge", "25%"), ("Problem Solving", "20%"),
        ("Communication Skills", "15%"), ("Cultural Fit", "15%"),
        ("Experience Relevance", "10%"), ("Leadership Potential", "10%"),
        ("Motivation & Drive", "5%"),
    ]
    for idx, (crit, weight) in enumerate(criteria):
        for col in range(1, max_col2 + 1):
            cell = ws2.cell(row=row, column=col)
            cell.font = FONT_NORMAL
            cell.fill = FILL_LIGHT_ORANGE if idx % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws2.cell(row=row, column=1).value = idx + 1
        ws2.cell(row=row, column=2).value = crit
        ws2.cell(row=row, column=8).value = weight
        row += 1

    add_footer(ws2, row, max_col2)
    add_print_setup(ws)
    add_print_setup(ws2)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_INTERVIEW_PROCESS_AND_CHECKLIST.xlsx"))
    print("  [OK] Interview Process & Checklist")


# ═══════════════════════════════════════════════════════════════════════════════
# 12. LEAVE REQUEST FORM
# ═══════════════════════════════════════════════════════════════════════════════
def create_leave_request():
    wb = Workbook()

    # Sheet 1: Leave Tracker
    ws = wb.active
    ws.title = "Leave Tracker"
    max_col = 14
    set_column_widths(ws, [5, 20, 18, 18, 14, 14, 14, 12, 14, 14, 14, 14, 14, 16])

    row = add_branded_header(ws, "LEAVE REQUEST TRACKER", "Track all employee leave", max_col)
    row = add_section_header(ws, row, "LEAVE REGISTER", max_col)

    headers = ["#", "Employee Name", "Employee ID", "Department",
               "Leave Type", "Start Date", "End Date", "Days",
               "Balance Before", "Balance After", "Relief Officer",
               "Line Manager", "HR Approved", "Status"]
    row = add_table_headers(ws, row, headers)

    dv_type = DataValidation(type="list",
                              formula1='"Annual,Sick,Compassionate,Maternity,Paternity,Study,Unpaid,Other"')
    ws.add_data_validation(dv_type)
    dv_status = DataValidation(type="list", formula1='"Pending,Approved,Denied,Cancelled,On Leave,Returned"')
    ws.add_data_validation(dv_status)

    for i in range(50):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv_type.add(ws.cell(row=row, column=5))
        dv_status.add(ws.cell(row=row, column=14))
        row += 1

    add_footer(ws, row, max_col)

    # Sheet 2: Fillable Leave Form
    ws2 = wb.create_sheet("Leave Form (Fillable)")
    max_col2 = 6
    set_column_widths(ws2, [4, 28, 30, 4, 28, 30])

    row = add_branded_header(ws2, "LEAVE REQUEST FORM", "Submit 7 days in advance", max_col2)

    row = add_section_header(ws2, row, "EMPLOYEE INFORMATION", max_col2)
    for label in ["Full Name:", "Employee ID:", "Department:", "Position:",
                   "Date of Joining:", "Phone (during leave):", "Email:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "LEAVE DETAILS", max_col2)
    for label in ["Leave Type:", "Start Date:", "End Date:",
                   "Working Days Requested:", "Resumption Date:", "Reason for Leave:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "LEAVE BALANCE", max_col2)
    for label in ["Annual Entitlement:", "Used to Date:", "Current Balance:", "Days Requested:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "HANDOVER ARRANGEMENTS", max_col2)
    for label in ["Relief Officer Name:", "Relief Officer Position:",
                   "Handover Notes Prepared:", "Pending Tasks Documented:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "APPROVAL CHAIN", max_col2)
    for label in ["Employee Signature:", "Date:", "Line Manager:", "LM Signature:", "LM Date:",
                   "HR Department:", "HR Signature:", "HR Date:",
                   "MD/CEO (if >5 days):", "CEO Signature:", "CEO Date:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)

    add_footer(ws2, row, max_col2)
    add_print_setup(ws)
    add_print_setup(ws2)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_LEAVE_REQUEST_FORM.xlsx"))
    print("  [OK] Leave Request Form")


# ═══════════════════════════════════════════════════════════════════════════════
# 13. STAFF ID CARD REQUEST
# ═══════════════════════════════════════════════════════════════════════════════
def create_staff_id():
    wb = Workbook()

    # Sheet 1: ID Card Tracker
    ws = wb.active
    ws.title = "ID Card Tracker"
    max_col = 12
    set_column_widths(ws, [5, 20, 18, 18, 14, 14, 18, 14, 14, 14, 16, 14])

    row = add_branded_header(ws, "STAFF ID CARD REQUEST TRACKER", "Track ID card requests", max_col)
    row = add_section_header(ws, row, "ID CARD REGISTER", max_col)

    headers = ["#", "Employee Name", "Employee ID", "Department",
               "Employment Type", "Request Type", "Request Date",
               "Photo Received", "Card Produced", "Card Issued",
               "Card Number", "Status"]
    row = add_table_headers(ws, row, headers)

    dv_type = DataValidation(type="list", formula1='"New Issue,Replacement"')
    ws.add_data_validation(dv_type)
    dv_emp = DataValidation(type="list", formula1='"Full-Time,Contract,Intern"')
    ws.add_data_validation(dv_emp)
    dv_status = DataValidation(type="list", formula1='"Pending,Processing,Ready,Issued,Cancelled"')
    ws.add_data_validation(dv_status)

    for i in range(40):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_LIGHT_ORANGE if i % 2 else FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = i + 1
        dv_type.add(ws.cell(row=row, column=6))
        dv_emp.add(ws.cell(row=row, column=5))
        dv_status.add(ws.cell(row=row, column=12))
        row += 1

    add_footer(ws, row, max_col)

    # Sheet 2: Fillable Form
    ws2 = wb.create_sheet("ID Request Form (Fillable)")
    max_col2 = 6
    set_column_widths(ws2, [4, 28, 30, 4, 28, 30])

    row = add_branded_header(ws2, "STAFF ID CARD REQUEST FORM", "Submit to HR Department", max_col2)

    row = add_section_header(ws2, row, "REQUEST DETAILS", max_col2)
    for label in ["Request Date:", "Request Type (New/Replace):"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "EMPLOYEE INFORMATION", max_col2)
    for label in ["Full Name:", "Employee ID:", "Department:", "Position:",
                   "Date of Employment:", "Employment Type:", "Office Location:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "REPLACEMENT DETAILS (if applicable)", max_col2)
    for label in ["Old Card Number:", "Reason for Replacement:", "Police Report Ref.:",
                   "Date of Loss/Damage:", "Replacement Fee Paid:", "Receipt Number:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "EMERGENCY CONTACTS", max_col2)
    for label in ["Primary Contact Name:", "Relationship:", "Phone:",
                   "Secondary Contact Name:", "Relationship:", "Phone:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)
    row += 1

    row = add_section_header(ws2, row, "APPROVAL", max_col2)
    for label in ["Employee Signature:", "Date:", "HR Officer:", "HR Signature:", "Date:"]:
        row = add_form_field(ws2, row, label, 2, 3, 1, max_col2)

    add_footer(ws2, row, max_col2)
    add_print_setup(ws)
    add_print_setup(ws2)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_STAFF_ID_CARD_REQUEST.xlsx"))
    print("  [OK] Staff ID Card Request")


# ═══════════════════════════════════════════════════════════════════════════════
# 14. JOB REQUIREMENTS HANDBOOK (Role Summary)
# ═══════════════════════════════════════════════════════════════════════════════
def create_job_requirements():
    wb = Workbook()
    ws = wb.active
    ws.title = "Roles & Requirements"
    max_col = 10
    set_column_widths(ws, [5, 14, 28, 18, 15, 15, 18, 14, 16, 14])

    row = add_branded_header(ws, "JOB REQUIREMENTS HANDBOOK", "All 37 roles across 6 departments", max_col)
    row = add_section_header(ws, row, "ROLES REGISTER", max_col)

    headers = ["#", "Role ID", "Job Title", "Department", "Reports To",
               "Location", "Employment Type", "Hiring Phase",
               "Compensation (NGN/mo)", "Status"]
    row = add_table_headers(ws, row, headers)

    dv_dept = DataValidation(type="list", formula1='"Engineering,Design,Product & Business,Operations,Support,Internship"')
    ws.add_data_validation(dv_dept)
    dv_type = DataValidation(type="list", formula1='"Full-Time + Equity,Full-Time,Contract,Internship"')
    ws.add_data_validation(dv_type)
    dv_phase = DataValidation(type="list", formula1='"Phase 1 (M1-4),Phase 2 (M5-8),Phase 3 (M9-12),Phase 4 (M13-16),Phase 5 (M17-20),Phase 6 (M21-24)"')
    ws.add_data_validation(dv_phase)
    dv_status = DataValidation(type="list", formula1='"Open,Interviewing,Offered,Filled,On Hold,Future"')
    ws.add_data_validation(dv_status)

    # Pre-fill with PECH roles
    roles = [
        ("PECH-ENG-CTO-001", "CTO / Founding Technical Architect", "Engineering", "CEO", "Full-Time + Equity", "Phase 1 (M1-4)", "800,000"),
        ("PECH-ENG-FS-001", "Full-Stack Platform Engineer", "Engineering", "CTO", "Full-Time", "Phase 1 (M1-4)", "500,000"),
        ("PECH-ENG-FS-002", "Full-Stack Platform Engineer #2", "Engineering", "CTO", "Full-Time", "Phase 2 (M5-8)", "500,000"),
        ("PECH-ENG-FE-001", "Frontend Engineer (React/Next.js)", "Engineering", "CTO", "Full-Time", "Phase 2 (M5-8)", "450,000"),
        ("PECH-ENG-BE-001", "Backend Engineer (Node.js/Python)", "Engineering", "CTO", "Full-Time", "Phase 2 (M5-8)", "450,000"),
        ("PECH-ENG-IOT-001", "IoT Engineer (Embedded/Firmware)", "Engineering", "CTO", "Full-Time", "Phase 1 (M1-4)", "450,000"),
        ("PECH-ENG-IOT-002", "IoT Field Engineer", "Engineering", "IoT Lead", "Full-Time", "Phase 3 (M9-12)", "350,000"),
        ("PECH-ENG-MOB-001", "Mobile Engineer (React Native)", "Engineering", "CTO", "Full-Time", "Phase 2 (M5-8)", "450,000"),
        ("PECH-ENG-DEV-001", "DevOps / Platform Engineer", "Engineering", "CTO", "Full-Time", "Phase 2 (M5-8)", "500,000"),
        ("PECH-ENG-DATA-001", "Senior Data Engineer / ML Ops", "Engineering", "CTO", "Full-Time", "Phase 3 (M9-12)", "500,000"),
        ("PECH-ENG-DATA-002", "Data Analyst", "Engineering", "Data Lead", "Full-Time", "Phase 3 (M9-12)", "350,000"),
        ("PECH-ENG-QA-001", "QA Engineer", "Engineering", "CTO", "Full-Time", "Phase 3 (M9-12)", "350,000"),
        ("PECH-DES-UI-001", "UI/UX Designer (Lead)", "Design", "CTO", "Full-Time", "Phase 1 (M1-4)", "400,000"),
        ("PECH-DES-UI-002", "UI/UX Designer #2", "Design", "Design Lead", "Full-Time", "Phase 3 (M9-12)", "350,000"),
        ("PECH-DES-GRA-001", "Graphic Designer / Brand", "Design", "Design Lead", "Full-Time", "Phase 2 (M5-8)", "300,000"),
        ("PECH-BIZ-PM-001", "Product Manager (Payments/PSSP)", "Product & Business", "CEO", "Full-Time", "Phase 2 (M5-8)", "600,000"),
        ("PECH-BIZ-BD-001", "Business Development Lead", "Product & Business", "CEO", "Full-Time", "Phase 2 (M5-8)", "500,000"),
        ("PECH-BIZ-MKT-001", "Market Operations Manager", "Product & Business", "BD Lead", "Full-Time", "Phase 3 (M9-12)", "400,000"),
        ("PECH-BIZ-COMP-001", "PSSP Compliance Officer", "Product & Business", "CEO", "Full-Time", "Phase 2 (M5-8)", "500,000"),
        ("PECH-OPS-LOG-001", "Logistics Coordinator", "Operations", "Ops Manager", "Full-Time", "Phase 3 (M9-12)", "300,000"),
        ("PECH-OPS-CS-001", "Customer Support Lead", "Operations", "Ops Manager", "Full-Time", "Phase 2 (M5-8)", "300,000"),
        ("PECH-OPS-CS-002", "Customer Support Rep", "Operations", "CS Lead", "Full-Time", "Phase 3 (M9-12)", "200,000"),
        ("PECH-OPS-CS-003", "Customer Support Rep #2", "Operations", "CS Lead", "Full-Time", "Phase 4 (M13-16)", "200,000"),
        ("PECH-OPS-FIN-001", "Finance & Accounts Officer", "Operations", "CEO", "Full-Time", "Phase 2 (M5-8)", "350,000"),
    ]

    for idx, (role_id, title, dept, reports, emp_type, phase, comp) in enumerate(roles):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_NORMAL
            cell.fill = FILL_LIGHT_ORANGE if idx % 2 else FILL_WHITE
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_LEFT
        ws.cell(row=row, column=1).value = idx + 1
        ws.cell(row=row, column=2).value = role_id
        ws.cell(row=row, column=3).value = title
        ws.cell(row=row, column=4).value = dept
        ws.cell(row=row, column=5).value = reports
        ws.cell(row=row, column=6).value = "Lagos, Nigeria"
        ws.cell(row=row, column=7).value = emp_type
        ws.cell(row=row, column=8).value = phase
        ws.cell(row=row, column=9).value = comp
        dv_status.add(ws.cell(row=row, column=10))
        row += 1

    # Add empty rows for additional roles
    for i in range(13):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_INPUT
            cell.fill = FILL_WHITE
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = len(roles) + i + 1
        dv_dept.add(ws.cell(row=row, column=4))
        dv_type.add(ws.cell(row=row, column=7))
        dv_phase.add(ws.cell(row=row, column=8))
        dv_status.add(ws.cell(row=row, column=10))
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_JOB_REQUIREMENTS_HANDBOOK.xlsx"))
    print("  [OK] Job Requirements Handbook")


# ═══════════════════════════════════════════════════════════════════════════════
# MASTER HR DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
def create_master_dashboard():
    """Create a master HR dashboard that summarizes all employment tracking."""
    wb = Workbook()
    ws = wb.active
    ws.title = "HR Dashboard"
    max_col = 8
    set_column_widths(ws, [5, 32, 18, 14, 14, 14, 14, 22])

    row = add_branded_header(ws, "PECH HR MASTER DASHBOARD", "Employment Documents Overview", max_col)

    row = add_section_header(ws, row, "DOCUMENT INDEX & STATUS", max_col)
    headers = ["#", "Document Name", "Category", "Total Records", "Active",
               "Pending", "Completed", "Google Sheets Link"]
    row = add_table_headers(ws, row, headers)

    docs = [
        ("Full-Time Employment Contract", "Contract"),
        ("Contract Worker Agreement", "Contract"),
        ("Internship Agreement", "Contract"),
        ("Office Marketer Agreement", "Contract"),
        ("Commission Marketer Agreement", "Contract"),
        ("Installer Agreement", "Contract"),
        ("API Developer Agreement", "Contract"),
        ("Guarantor / Surety Form", "Form"),
        ("Non-Disclosure Agreement", "Contract"),
        ("Candidate Application Form", "HR Form"),
        ("Interview Process & Checklist", "HR Form"),
        ("Leave Request Form", "HR Form"),
        ("Staff ID Card Request", "HR Form"),
        ("Job Requirements Handbook", "Reference"),
    ]

    for idx, (name, cat) in enumerate(docs):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_NORMAL
            cell.fill = FILL_LIGHT_ORANGE if idx % 2 else FILL_WHITE
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_LEFT
        ws.cell(row=row, column=1).value = idx + 1
        ws.cell(row=row, column=2).value = name
        ws.cell(row=row, column=3).value = cat
        ws.cell(row=row, column=8).value = "(paste Google Sheets URL here)"
        ws.cell(row=row, column=8).font = FONT_INPUT
        row += 1

    row += 2
    row = add_section_header(ws, row, "TEAM SUMMARY (37 Roles: 24 FT + 5 Contract + 8 Intern)", max_col)
    summary_headers = ["#", "Category", "Target Count", "Hired", "In Progress", "Open", "% Complete", "Notes"]
    row = add_table_headers(ws, row, summary_headers)

    categories = [
        ("Full-Time Employees", "24"), ("Contract Workers", "5"),
        ("Interns", "8"), ("TOTAL", "37"),
    ]
    for idx, (cat, target) in enumerate(categories):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=10, bold=(cat == "TOTAL"), color="333333")
            cell.fill = FILL_LIGHT_BLUE if cat == "TOTAL" else (FILL_LIGHT_ORANGE if idx % 2 else FILL_WHITE)
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1).value = idx + 1
        ws.cell(row=row, column=2).value = cat
        ws.cell(row=row, column=3).value = target
        row += 1

    row += 2
    row = add_section_header(ws, row, "GOOGLE SHEETS SETUP INSTRUCTIONS", max_col)
    instructions = [
        "1. Open Google Drive (drive.google.com)",
        "2. Click '+ New' > 'File upload' > Select the .xlsx file",
        "3. Double-click the uploaded file to open in Google Sheets",
        "4. Google Sheets will automatically convert the file",
        "5. Click 'File' > 'Save as Google Sheets' to create native version",
        "6. Share via 'Share' button (top-right) with team members",
        "7. Set permissions: 'Editor' for HR, 'Viewer' for managers",
        "8. Copy the sharing URL and paste in the 'Google Sheets Link' column above",
        "",
        "TIP: All dropdown validations and formatting will be preserved in Google Sheets!",
        "TIP: Use 'File > Make a copy' to create templates for new hires.",
    ]
    for instr in instructions:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        cell = ws.cell(row=row, column=1)
        cell.value = f"  {instr}"
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = 20
        row += 1

    add_footer(ws, row, max_col)
    add_print_setup(ws)
    wb.save(os.path.join(OUTPUT_DIR, "PECH_HR_MASTER_DASHBOARD.xlsx"))
    print("  [OK] HR Master Dashboard")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    ensure_output_dir()

    print("\n" + "=" * 70)
    print("  PECH Group Holdings Ltd — Employment Excel Generator")
    print("  Generating branded .xlsx files for all employment documents...")
    print("=" * 70 + "\n")

    # Contracts
    create_fulltime_contract()
    create_contract_worker()
    create_internship()
    create_office_marketer()
    create_commission_marketer()
    create_installer()
    create_guarantor_form()
    create_nda()
    create_api_developer()

    # HR Forms
    create_candidate_application()
    create_interview_checklist()
    create_leave_request()
    create_staff_id()
    create_job_requirements()

    # Master Dashboard
    create_master_dashboard()

    print("\n" + "=" * 70)
    print(f"  SUCCESS! 15 Excel files generated in: {OUTPUT_DIR}/")
    print("=" * 70)
    print("\n  Files created:")
    for f in sorted(os.listdir(OUTPUT_DIR)):
        if f.endswith(".xlsx"):
            size = os.path.getsize(os.path.join(OUTPUT_DIR, f))
            print(f"    - {f} ({size / 1024:.1f} KB)")

    print("\n  GOOGLE SHEETS: Upload any .xlsx to Google Drive, then")
    print("  open with Google Sheets. All formatting is preserved.")
    print("  See PECH_HR_MASTER_DASHBOARD.xlsx for full instructions.\n")


if __name__ == "__main__":
    main()
